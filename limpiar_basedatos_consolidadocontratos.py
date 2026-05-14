"""
pivot_to_consolidado.py
========================
Toma el archivo Pivot descargado desde SAP Ariba y lo integra al
Consolidado de Contratos, actualizando la hoja "Info Ariba" con los
nuevos datos, manteniendo intactas las hojas BG, Antiguo y las fórmulas
del Consolidado principal.
 
Uso:
    python pivot_to_consolidado.py <pivot.xlsx> <consolidado.xlsx> [salida.xlsx]
 
Si no se indica salida, sobreescribe el consolidado.
"""
 
import sys
import shutil
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
 
# ─────────────────────────────────────────────────────────────
# COLUMNAS ESPERADAS EN Info Ariba (igual que en el Pivot Data)
# La col A queda vacía (None), los datos van desde col B
# ─────────────────────────────────────────────────────────────
INFO_ARIBA_HEADERS = [
    None,                                               # col A – siempre vacía
    'ID de contrato',
    'Proyecto - Nombre del proyecto',
    'Fecha de inicio',
    'Nombre del propietario',
    'Código acreedor SAP',
    'Es Indefinido',
    'Región - Región (L2)',
    'Rut empresa proveedor',
    'Partes afectadas - Proveedor común',
    'Contrato - Contrato',
    'Fecha de entrada en vigor - Fecha',
    'Fecha de finalización - Año',
    'Estado del contrato',
    'Fecha de expiración - Fecha',
    'Es un proyecto de prueba',
    'Descripción',
    'Aplica Garantía',
    'Fecha de presentación Garantía N°1 - Fecha',
    'Fecha de termino de notificaciones de garantía - Fecha',
    'N° de Tipos de Garantías',
    'Fecha de termino de notificaciones de garantía - Año',
    'sum(Importe del contrato)',
    'sum(Importe Monto Total Contrato Original)',
    'sum(Importe Monto total Contrato)',
    'Sample',
]
 
# Columnas pivot → nombres esperados en Info Ariba (para renombrar si difieren)
PIVOT_RENAME = {
    'Fecha de termino de notificaciones de garantía - Fecha': 'Fecha de termino de notificaciones de garantía - Fecha',
}
 
 
def find_header_row(xl_path: Path, sheet: str) -> int:
    """Encuentra la fila (0-indexed para pandas) donde está 'ID de contrato'."""
    df = pd.read_excel(xl_path, sheet_name=sheet, header=None, nrows=30, engine='openpyxl')
    for i, row in df.iterrows():
        if 'ID de contrato' in row.values:
            return i
    raise ValueError(f"No se encontró 'ID de contrato' en la hoja '{sheet}' de {xl_path.name}")
 
 
def read_pivot(pivot_path: Path) -> pd.DataFrame:
    """Lee la hoja Data del Pivot y retorna un DataFrame limpio."""
    header_idx = find_header_row(pivot_path, 'Data')
    df = pd.read_excel(pivot_path, sheet_name='Data', header=header_idx, engine='openpyxl')
    df = df.dropna(how='all').dropna(axis=1, how='all')
    # Renombrar columnas si hace falta
    df = df.rename(columns=PIVOT_RENAME)
    return df
 
 
def clear_info_ariba(ws):
    """Borra todas las filas de datos de Info Ariba (deja solo la fila 1 de header)."""
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
 
 
def write_info_ariba(ws, df: pd.DataFrame):
    """
    Escribe los datos del pivot en la hoja Info Ariba.
    Estructura: col A vacía, col B = 'ID de contrato', etc.
    """
    # Asegurar que los headers estén en fila 1
    for col_idx, header in enumerate(INFO_ARIBA_HEADERS, start=1):
        ws.cell(row=1, column=col_idx).value = header
 
    # Escribir datos fila a fila desde fila 2
    pivot_cols = [h for h in INFO_ARIBA_HEADERS if h is not None]
 
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        ws.cell(row=row_idx, column=1).value = None   # col A siempre vacía
        for col_offset, col_name in enumerate(pivot_cols, start=2):
            val = row.get(col_name, None)
            # Convertir NaN / NaT a None
            try:
                if pd.isna(val):
                    val = None
            except (TypeError, ValueError):
                pass
            ws.cell(row=row_idx, column=col_offset).value = val
 
 
def update_consolidado_new_rows(ws_cons, ws_ia, existing_ids: set):
    """
    Agrega al Consolidado de Contratos las filas del pivot que son nuevas
    (no existen en la hoja Antiguo ni en el consolidado actual).
    Copia las fórmulas del último row existente adaptando el número de fila.
    """
    # Encontrar última fila con ID real (col A no vacía)
    last_data_row = 1
    for r in range(2, ws_cons.max_row + 1):
        if ws_cons.cell(row=r, column=1).value:
            last_data_row = r
 
    # IDs ya en el consolidado
    cons_ids = {ws_cons.cell(row=r, column=1).value for r in range(2, ws_cons.max_row + 1)}
 
    # IDs en Info Ariba
    new_ids = []
    for r in range(2, ws_ia.max_row + 1):
        cw_id = ws_ia.cell(row=r, column=2).value
        if cw_id and cw_id not in cons_ids and cw_id not in existing_ids:
            new_ids.append(cw_id)
 
    if not new_ids:
        return 0
 
    # Fórmulas plantilla de la última fila con datos reales
    template_row = last_data_row
    template_formulas = {}
    for col in range(1, ws_cons.max_column + 1):
        template_formulas[col] = ws_cons.cell(row=template_row, column=col).value
 
    # Estilos del header
    thin = Border(
        left=Side('hair'), right=Side('hair'),
        top=Side('hair'), bottom=Side('hair')
    )
    data_font = Font(name='Arial', size=8)
    data_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
 
    next_row = last_data_row + 1
    for new_id in new_ids:
        new_row = next_row
        next_row += 1
        for col in range(1, ws_cons.max_column + 1):
            cell = ws_cons.cell(row=new_row, column=col)
            if col == 1:
                cell.value = new_id
            else:
                tmpl = template_formulas.get(col)
                if isinstance(tmpl, str) and tmpl.startswith('='):
                    # Reemplazar número de fila de referencia en la fórmula
                    updated = tmpl.replace(str(template_row), str(new_row))
                    cell.value = updated
                else:
                    cell.value = None
            cell.font = data_font
            cell.border = thin
            cell.alignment = data_align
 
    return len(new_ids)
 
 
def apply_header_style(ws):
    """Aplica el estilo del header al consolidado (fila 1)."""
    hdr_fill = PatternFill('solid', fgColor='FFE7E6E6')
    hdr_font = Font(name='Arial', size=8, bold=True)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(
        left=Side('hair'), right=Side('hair'),
        top=Side('hair'), bottom=Side('hair')
    )
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = thin
    ws.row_dimensions[1].height = 45
 
 
def main(pivot_path: str, consolidado_path: str, output_path: str = None):
    pivot_path = Path(pivot_path)
    consolidado_path = Path(consolidado_path)
    output_path = Path(output_path) if output_path else consolidado_path
 
    if not pivot_path.exists():
        raise FileNotFoundError(f"No se encontró el pivot: {pivot_path}")
    if not consolidado_path.exists():
        raise FileNotFoundError(f"No se encontró el consolidado: {consolidado_path}")
 
    # Copiar el consolidado al destino si es diferente
    if output_path != consolidado_path:
        shutil.copy2(consolidado_path, output_path)
 
    print(f"📂 Leyendo pivot: {pivot_path.name}")
    df_pivot = read_pivot(pivot_path)
    print(f"   → {len(df_pivot)} contratos encontrados en el pivot")
 
    print(f"📂 Abriendo consolidado: {output_path.name}")
    wb = openpyxl.load_workbook(output_path)
 
    ws_ia = wb['Info Ariba']
    ws_cons = wb['Consolidado de Contratos']
    ws_ant = wb['Antiguo']
 
    # IDs que ya existen en Antiguo
    existing_ids = {ws_ant.cell(row=r, column=1).value for r in range(2, ws_ant.max_row + 1)}
 
    print("🔄 Actualizando hoja Info Ariba con datos del pivot...")
    clear_info_ariba(ws_ia)
    write_info_ariba(ws_ia, df_pivot)
    print(f"   → {len(df_pivot)} filas escritas en Info Ariba")
 
    print("🔄 Verificando contratos nuevos para agregar al Consolidado...")
    added = update_consolidado_new_rows(ws_cons, ws_ia, existing_ids)
    if added:
        print(f"   → {added} contratos nuevos agregados al Consolidado")
    else:
        print("   → No hay contratos nuevos que agregar")
 
    # Re-aplicar estilos del header por si acaso
    apply_header_style(ws_cons)
 
    # Mantener freeze y filtros
    ws_cons.freeze_panes = 'E3'
    ws_cons.auto_filter.ref = ws_cons.dimensions
 
    wb.save(output_path)
    print(f"\n✅ Consolidado actualizado guardado en: {output_path.resolve()}")
    return str(output_path)
 
 
if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: python pivot_to_consolidado.py <pivot.xlsx> <consolidado.xlsx> [salida.xlsx]")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else None)