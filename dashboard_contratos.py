"""
dashboard_pivot.py — Softys Chile · Compras Estratégicas
=========================================================
VERSIÓN CON PERSISTENCIA + SINCRONIZACIÓN ONEDRIVE
====================================================
El Pivot de Ariba y el Consolidado se guardan en ./data/pivot_persistente/
El Consolidado puede vincularse a un link de OneDrive/SharePoint para
actualizarse con un clic desde la interfaz.

Instalar:  pip install streamlit pandas plotly openpyxl requests
Ejecutar:  streamlit run dashboard_pivot.py
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
from io import BytesIO
import hashlib
import openpyxl
import warnings
import json
import os
import re
import shutil
warnings.filterwarnings("ignore")

from config import (
    ESTRATEGICOS, TACTICOS, TODOS_COMPRADORES,
    PIVOT_A_CANON, PIVOT_COL_MAP, CONS_COL_MAP,
    RIESGO_COLORES, SYNC_COLORES, SYNC_BG, SYNC_FG,
)

# ──────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ──────────────────────────────────────────────────────────────
st.set_page_config(page_title="Gestión de Contratos · Softys", page_icon="📋",
                   layout="wide", initial_sidebar_state="expanded")

# ══════════════════════════════════════════════════════════════
# BLOQUE 1: PERSISTENCIA DE ARCHIVOS
# ══════════════════════════════════════════════════════════════

PERSIST_DIR   = os.path.join(os.path.dirname(__file__), "data", "pivot_persistente")
PERSIST_PIVOT = os.path.join(PERSIST_DIR, "pivot_ariba.xlsx")
PERSIST_CONS  = os.path.join(PERSIST_DIR, "consolidado_drive.xlsx")
PERSIST_META  = os.path.join(PERSIST_DIR, "metadata.json")
PERSIST_OD    = os.path.join(PERSIST_DIR, "onedrive_config.json")

os.makedirs(PERSIST_DIR, exist_ok=True)


def _md5(data: bytes) -> str:
    return hashlib.md5(data).hexdigest()


def _fmt_size(size_bytes: int) -> str:
    if size_bytes >= 1_048_576:
        return f"{size_bytes/1_048_576:.1f} MB"
    if size_bytes >= 1_024:
        return f"{size_bytes/1_024:.0f} KB"
    return f"{size_bytes} B"


def _session_id() -> str:
    if "session_id" not in st.session_state:
        import uuid
        st.session_state["session_id"] = str(uuid.uuid4())[:8]
    return st.session_state["session_id"]


# ── Metadata general ─────────────────────────────────────────

def cargar_metadata() -> dict:
    if not os.path.exists(PERSIST_META):
        return {}
    try:
        with open(PERSIST_META, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _escribir_metadata(meta: dict):
    with open(PERSIST_META, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)


# ── Pivot ─────────────────────────────────────────────────────

def guardar_pivot_persistente(content: bytes, filename: str, session_id: str = "desconocido") -> dict:
    tmp_path = PERSIST_PIVOT + ".tmp"
    try:
        with open(tmp_path, "wb") as f:
            f.write(content)
        shutil.move(tmp_path, PERSIST_PIVOT)
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
    meta = cargar_metadata()
    meta["pivot"] = {
        "filename":    filename,
        "hash_md5":    _md5(content),
        "size_bytes":  len(content),
        "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "uploaded_by": session_id,
    }
    _escribir_metadata(meta)
    cargar_pivot.clear()
    return meta


def eliminar_pivot_persistente():
    if os.path.exists(PERSIST_PIVOT):
        os.remove(PERSIST_PIVOT)
    meta = cargar_metadata()
    meta.pop("pivot", None)
    _escribir_metadata(meta)
    cargar_pivot.clear()


def leer_pivot_persistente() -> tuple:
    meta = cargar_metadata()
    meta_pivot = meta.get("pivot")
    if not meta_pivot or not os.path.exists(PERSIST_PIVOT):
        return None, None
    try:
        with open(PERSIST_PIVOT, "rb") as f:
            content = f.read()
        if _md5(content) != meta_pivot.get("hash_md5", ""):
            st.warning("⚠️ El archivo Pivot guardado no supera la validación de integridad (MD5 distinto). Por favor, vuelve a subirlo.")
            return None, None
        return content, meta_pivot
    except Exception as e:
        st.warning(f"⚠️ Error al leer el Pivot guardado: {e}. Por favor, vuelve a subirlo.")
        return None, None


# ── Consolidado ───────────────────────────────────────────────

def guardar_cons_persistente(content: bytes, filename: str, session_id: str = "desconocido") -> dict:
    tmp_path = PERSIST_CONS + ".tmp"
    try:
        with open(tmp_path, "wb") as f:
            f.write(content)
        shutil.move(tmp_path, PERSIST_CONS)
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
    meta = cargar_metadata()
    meta["consolidado"] = {
        "filename":    filename,
        "hash_md5":    _md5(content),
        "size_bytes":  len(content),
        "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "uploaded_by": session_id,
    }
    _escribir_metadata(meta)
    cargar_consolidado.clear()
    return meta


def eliminar_cons_persistente():
    if os.path.exists(PERSIST_CONS):
        os.remove(PERSIST_CONS)
    meta = cargar_metadata()
    meta.pop("consolidado", None)
    _escribir_metadata(meta)
    cargar_consolidado.clear()


def leer_cons_persistente() -> tuple:
    meta = cargar_metadata()
    meta_cons = meta.get("consolidado")
    if not meta_cons or not os.path.exists(PERSIST_CONS):
        return None, None
    try:
        with open(PERSIST_CONS, "rb") as f:
            content = f.read()
        if _md5(content) != meta_cons.get("hash_md5", ""):
            st.warning("⚠️ El Consolidado guardado no supera la validación de integridad. Por favor, vuelve a subirlo.")
            return None, None
        return content, meta_cons
    except Exception as e:
        st.warning(f"⚠️ Error al leer el Consolidado guardado: {e}.")
        return None, None


# ══════════════════════════════════════════════════════════════
# BLOQUE 2: FUNCIONES ONEDRIVE
# ══════════════════════════════════════════════════════════════

def _leer_od_config() -> dict:
    if not os.path.exists(PERSIST_OD):
        return {}
    try:
        with open(PERSIST_OD, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _guardar_od_config(cfg: dict):
    with open(PERSIST_OD, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def guardar_url_onedrive(url: str) -> None:
    cfg = _leer_od_config()
    cfg["url"]        = url.strip()
    cfg["guardado_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _guardar_od_config(cfg)


def eliminar_url_onedrive() -> None:
    if os.path.exists(PERSIST_OD):
        os.remove(PERSIST_OD)


def _transformar_url_onedrive(url: str) -> str | None:
    """
    Convierte un link de compartición OneDrive/SharePoint en URL de descarga directa.
    - OneDrive personal (1drv.ms, onedrive.live.com): añade download=1
    - SharePoint corporativo (:x:, :xlsx:, /r/): añade download=1
    - URLs que ya terminan en /content o tienen download=1: las devuelve tal cual
    """
    url = url.strip()
    if not url:
        return None

    # Ya tiene parámetro de descarga directa
    if "download=1" in url:
        return url

    # Link corto 1drv.ms — requests seguirá el redirect automáticamente
    if "1drv.ms" in url:
        sep = "&" if "?" in url else "?"
        return url + sep + "download=1"

    # OneDrive personal live.com
    if "onedrive.live.com" in url:
        sep = "&" if "?" in url else "?"
        return url + sep + "download=1"

    # SharePoint / OneDrive for Business (xxx.sharepoint.com)
    if "sharepoint.com" in url:
        sep = "&" if "?" in url else "?"
        return url + sep + "download=1"

    # Fallback: intentar añadir download=1 de todas formas
    sep = "&" if "?" in url else "?"
    return url + sep + "download=1"


def descargar_consolidado_onedrive(url: str) -> tuple:
    """
    Descarga el Excel desde OneDrive/SharePoint.
    Retorna (bytes, "") si tiene éxito o (None, mensaje_error) si falla.
    """
    try:
        import requests
    except ImportError:
        return None, "❌ Falta la librería 'requests'. Ejecuta: pip install requests"

    url_dl = _transformar_url_onedrive(url)
    if not url_dl:
        return None, "❌ URL vacía o no reconocida."

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
    }

    try:
        resp = requests.get(url_dl, headers=headers, allow_redirects=True, timeout=45)
    except requests.exceptions.Timeout:
        return None, "❌ Tiempo de espera agotado (>45s). Verifica la URL o tu conexión a internet."
    except requests.exceptions.ConnectionError as e:
        return None, f"❌ Error de conexión: {e}"
    except Exception as e:
        return None, f"❌ Error inesperado al conectar: {e}"

    if resp.status_code == 401:
        return None, (
            "🔐 Acceso denegado (401 — sin autorización).\n\n"
            "El archivo requiere login corporativo. Debes generar un link con permiso "
            "**'Cualquier persona con el link puede ver'** desde SharePoint:\n\n"
            "1. Abre el archivo en OneDrive/SharePoint\n"
            "2. Clic en **Compartir** → **Personas con el link**\n"
            "3. Cambia a **'Cualquier persona con el link'**\n"
            "4. Copia el link y pégalo aquí"
        )
    if resp.status_code == 403:
        return None, (
            "🔐 Acceso prohibido (403).\n\n"
            "Verifica que el link sea de tipo **'Cualquier persona con el link puede ver'**. "
            "Los links de tipo 'Solo personas de tu organización' requieren login y no funcionan aquí."
        )
    if resp.status_code == 404:
        return None, "❌ Archivo no encontrado (404). El archivo puede haber sido movido o eliminado. Verifica el link."
    if resp.status_code != 200:
        return None, f"❌ Error HTTP {resp.status_code} al descargar. Verifica el link e intenta de nuevo."

    # Detectar si el servidor devolvió una página de login HTML en lugar del archivo
    content_type = resp.headers.get("Content-Type", "")
    if ("html" in content_type.lower() and
            len(resp.content) > 0 and
            resp.content[:15].lower().startswith((b"<!doctype", b"<html"))):
        return None, (
            "🔐 El servidor devolvió una página de login en lugar del archivo.\n\n"
            "Esto ocurre cuando el link requiere autenticación corporativa. "
            "Genera un link **'Cualquier persona con el link puede ver'** desde SharePoint."
        )

    if len(resp.content) < 500:
        return None, "❌ El archivo descargado parece vacío o incompleto. Verifica el link."

    # Validar magic bytes de Excel
    magic = resp.content[:4]
    is_xlsx = magic == b'PK\x03\x04'       # xlsx / zip
    is_xls  = magic == b'\xd0\xcf\x11\xe0' # xls / OLE2
    if not (is_xlsx or is_xls):
        return None, (
            "❌ El archivo descargado no es un Excel válido.\n\n"
            "Asegúrate de que el link apunte directamente al archivo .xlsx, "
            "no a una carpeta ni a la vista previa del navegador."
        )

    return resp.content, ""


def sincronizar_desde_onedrive(url: str, session_id: str) -> tuple:
    """
    Descarga y persiste el Consolidado desde OneDrive.
    Retorna (éxito: bool, mensaje: str).
    """
    content, err = descargar_consolidado_onedrive(url)
    if not content:
        return False, err

    # Intentar extraer nombre del archivo del URL
    match = re.search(r'/([^/?#]+\.xlsx)', url, re.IGNORECASE)
    nombre = match.group(1) if match else "consolidado_onedrive.xlsx"

    try:
        guardar_cons_persistente(content, nombre, session_id)
        # Actualizar registro de último sync exitoso
        cfg = _leer_od_config()
        cfg["ultimo_sync_ok"]   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cfg["ultimo_sync_size"] = len(content)
        _guardar_od_config(cfg)
        return True, f"✅ Consolidado actualizado desde OneDrive ({_fmt_size(len(content))})"
    except Exception as e:
        return False, f"❌ Error al guardar el archivo descargado: {e}"


# ──────────────────────────────────────────────────────────────
# PALETA Y ESTILOS SOFTYS
# ──────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
}

/* ── SIDEBAR: fondo oscuro → texto BLANCO (selectivo) ── */
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #003F7A 0%, #005CA9 60%, #0072CE 100%);
    border-right: 1px solid #004f96;
}

/* Texto blanco para labels, títulos y elementos directos de la sidebar */
section[data-testid="stSidebar"] .stMarkdown p,
section[data-testid="stSidebar"] .stMarkdown h1,
section[data-testid="stSidebar"] .stMarkdown h2,
section[data-testid="stSidebar"] .stMarkdown h3,
section[data-testid="stSidebar"] label:not([class*="FileUploader"]) span,
section[data-testid="stSidebar"] > div > div > label,
section[data-testid="stSidebar"] .stCheckbox label span {
    color: #FFFFFF !important;
}

/* Selectboxes y multiselect en sidebar - fondo semi-transparente, texto blanco */
section[data-testid="stSidebar"] .stSelectbox > div > div,
section[data-testid="stSidebar"] .stMultiSelect > div > div {
    background: rgba(255,255,255,0.15) !important;
    border-color: rgba(255,255,255,0.35) !important;
    color: #FFFFFF !important;
    border-radius: 8px !important;
}

/* File uploaders en sidebar - fondo blanco, texto negro */
section[data-testid="stSidebar"] .stFileUploader,
section[data-testid="stSidebar"] .stFileUploader *,
section[data-testid="stSidebar"] [class*="stFileUploader"] *,
section[data-testid="stSidebar"] div[data-testid="stFileUploader"] * {
    color: #000000 !important;
}

/* Alerts/cards con fondo claro en sidebar - texto oscuro */
section[data-testid="stSidebar"] .stAlert *,
section[data-testid="stSidebar"] .persist-card *,
section[data-testid="stSidebar"] div[data-testid="stExpander"] * {
    color: #000000 !important;
}

/* HR y captions en sidebar */
section[data-testid="stSidebar"] hr { 
    border-color: rgba(255,255,255,0.25); 
    margin: 10px 0; 
}
section[data-testid="stSidebar"] .stCaption { 
    color: rgba(255,255,255,0.85) !important; 
    font-size:.72rem !important; 
}

/* Botón colapsable de sidebar */
button[kind="headerNoPadding"], [data-testid="collapsedControl"] {
    display: flex !important; visibility: visible !important; opacity: 1 !important;
    background: #005CA9 !important; border-radius: 0 8px 8px 0 !important;
    border: 1px solid #0072CE !important; border-left: none !important;
    padding: 6px 4px !important; z-index: 999999 !important;
}
[data-testid="collapsedControl"] svg { fill: #FFFFFF !important; color: #FFFFFF !important; visibility: visible !important; opacity: 1 !important; }
[data-testid="collapsedControl"] { position: fixed !important; top: 50% !important; left: 0 !important; transform: translateY(-50%) !important; }

/* ── CONTENIDO PRINCIPAL: fondo claro → texto OSCURO ── */
.stApp { background: #F5F7FA; }
.block-container { padding-top: 1.4rem; padding-bottom: 2rem; }

/* Tabs */
.stTabs [data-baseweb="tab-list"] { background: transparent; gap: 4px; border-bottom: 2px solid #D1DCE8; padding-bottom: 0; }
.stTabs [data-baseweb="tab"] { background: transparent; border: none; border-bottom: 3px solid transparent; color: #1A2E44; font-size: .82rem; font-weight: 600; padding: 8px 16px; letter-spacing: .02em; border-radius: 6px 6px 0 0; transition: all .15s ease; }
.stTabs [data-baseweb="tab"]:hover { color: #005CA9; background: #EAF2FB; }
.stTabs [aria-selected="true"] { color: #005CA9 !important; border-bottom: 3px solid #005CA9 !important; background: transparent !important; }
.stTabs [data-baseweb="tab-panel"] { padding-top: 14px; }

/* KPI Cards */
.kpi-row { display: grid; grid-template-columns: repeat(5, 1fr); gap: 12px; margin-bottom: 16px; }
.kpi { background: #ffffff; border-radius: 12px; padding: 16px 14px 13px; border-top: 3px solid #0072CE; box-shadow: 0 1px 4px rgba(0,92,169,.08), 0 4px 12px rgba(0,0,0,.04); transition: box-shadow .2s ease; }
.kpi:hover { box-shadow: 0 4px 16px rgba(0,92,169,.13); }
.kpi.g  { border-top-color: #00A651; }
.kpi.r  { border-top-color: #E02020; }
.kpi.y  { border-top-color: #F59E0B; }
.kpi.gr { border-top-color: #8FA3B8; }
.kpi.b  { border-top-color: #0072CE; }
.kpi.p  { border-top-color: #6C3FC4; }
.kpi.o  { border-top-color: #EA580C; }
.kpi-lbl { font-size: .65rem; text-transform: uppercase; letter-spacing: .08em; color: #5A7490; font-weight: 700; margin-bottom: 4px; }
.kpi-val { font-size: 1.9rem; font-weight: 800; color: #000000; line-height: 1.05; letter-spacing: -.02em; }
.kpi-sub { font-size: .68rem; color: #4A5568; margin-top: 3px; font-weight: 400; }

/* Alert Cards */
.alert-card { border-radius: 10px; padding: 12px 16px; margin-bottom: 10px; font-size: .84rem; line-height: 1.6; }
.alert-card.red    { background: #FEF2F2; border-left: 4px solid #E02020; color: #7F1D1D; }
.alert-card.yellow { background: #FFFBEB; border-left: 4px solid #F59E0B; color: #78350F; }
.alert-card.blue   { background: #EAF2FB; border-left: 4px solid #0072CE; color: #1E3A5F; }
.alert-card.green  { background: #F0FDF4; border-left: 4px solid #00A651; color: #14532D; }
.alert-card.purple { background: #F5F0FF; border-left: 4px solid #6C3FC4; color: #3B1A78; }

/* Persist Card */
.persist-card {
    background: #ffffff;
    border-radius: 12px;
    border-left: 4px solid #00A651;
    padding: 12px 16px;
    margin: 8px 0 10px;
    box-shadow: 0 1px 4px rgba(0,0,0,.06);
}
.persist-card .pc-label {
    font-size: .62rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: .08em;
    color: #00A651;
    margin-bottom: 3px;
}
.persist-card .pc-filename {
    font-size: .82rem;
    font-weight: 700;
    color: #000000;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}
.persist-card .pc-meta {
    font-size: .68rem;
    color: #4A5568;
    margin-top: 2px;
    line-height: 1.5;
}

/* Section headers */
.sec { font-size: .9rem; font-weight: 700; color: #005CA9; border-bottom: 2px solid #D1DCE8; padding-bottom: 6px; margin: 22px 0 12px; letter-spacing: .01em; display: flex; align-items: center; gap: 6px; }

/* DataFrames */
[data-testid="stDataFrame"] { border-radius: 10px; overflow: hidden; border: 1px solid #D1DCE8; box-shadow: 0 1px 4px rgba(0,0,0,.04); }

/* Download buttons */
.stDownloadButton > button { background: #005CA9; color: #ffffff; border: none; border-radius: 8px; font-size: .78rem; font-weight: 600; padding: 8px 14px; width: 100%; transition: background .15s ease; }
.stDownloadButton > button:hover { background: #003F7A; color: #ffffff; }

/* Expanders */
[data-testid="stExpander"] { border: 1px solid #D1DCE8 !important; border-radius: 10px !important; background: #ffffff; margin-bottom: 8px; }
[data-testid="stExpander"] summary { font-weight: 600; font-size: .86rem; color: #000000; padding: 10px 14px; }
[data-testid="stExpander"] summary:hover { background: #F0F6FF; }

/* Badges - texto oscuro sobre fondo claro */
.badge-ariba  { background:#E8F2FB; color:#003F7A; border-radius:99px; padding:2px 9px; font-size:.67rem; font-weight:700; }
.badge-cons   { background:#F0EAFF; color:#3B1A78; border-radius:99px; padding:2px 9px; font-size:.67rem; font-weight:700; }
.badge-ambos  { background:#E6F9EE; color:#005F2A; border-radius:99px; padding:2px 9px; font-size:.67rem; font-weight:700; }

.stSpinner > div { border-top-color: #005CA9 !important; }

#MainMenu, footer, header { visibility: hidden; }
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────
# UTILIDADES
# ──────────────────────────────────────────────────────────────
def norm(v) -> str:
    return str(v).strip().lower() if pd.notna(v) and str(v).strip() not in ("", "nan") else ""

def canon(raw: str) -> str:
    k = norm(raw)
    return PIVOT_A_CANON.get(k, str(raw).strip() if k else "Sin asignar")

def es_comprador_oficial(nombre: str) -> bool:
    return nombre in TODOS_COMPRADORES

def tipo_comprador(nombre: str) -> str:
    es_e = nombre in ESTRATEGICOS
    es_t = nombre in TACTICOS
    if es_e and es_t: return "Estratégico + Táctico"
    if es_e: return "Estratégico"
    if es_t: return "Táctico"
    return "No registrado"

def parse_fecha(v) -> pd.Timestamp:
    if pd.isna(v): return pd.NaT
    s = str(v).strip()
    if s in ("", "99.99.9999", "31/12/2999", "2999", "nan"): return pd.NaT
    if isinstance(v, pd.Timestamp):
        return v if v.year < 2900 else pd.NaT
    if isinstance(v, (int, float)):
        try:
            ts = pd.Timestamp("1899-12-30") + pd.Timedelta(days=int(v))
            return ts if ts.year < 2900 else pd.NaT
        except: return pd.NaT
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            ts = pd.to_datetime(s, format=fmt)
            return ts if ts.year < 2900 else pd.NaT
        except: continue
    try:
        ts = pd.to_datetime(s, dayfirst=True, errors="coerce")
        return ts if pd.notna(ts) and ts.year < 2900 else pd.NaT
    except: return pd.NaT

def calcular_riesgo(estado: str, dias, es_indef: bool) -> str:
    if es_indef: return "BAJO 🟢"
    if dias is None or pd.isna(dias): return "REVISAR ⚪"
    d = int(dias)
    if estado in ("Vencido", "Cancelado", "Terminado") or d < 0: return "ALTO 🔴"
    if estado in ("Modificación del borrador", "Próximo a vencer") or d <= 60: return "MEDIO 🟡"
    if estado in ("Borrador", "En espera"): return "REVISAR ⚪"
    if d > 60: return "BAJO 🟢"
    return "REVISAR ⚪"

def fmt_m(v: float) -> str:
    if v >= 1_000_000_000: return f"${v/1_000_000_000:.1f}B"
    if v >= 1_000_000:     return f"${v/1_000_000:.1f}M"
    if v >= 1_000:         return f"${v/1_000:.0f}K"
    return f"${v:.0f}"


# ──────────────────────────────────────────────────────────────
# CARGA Y TRANSFORMACIÓN
# ──────────────────────────────────────────────────────────────
def detectar_header(content: bytes) -> int:
    df_scan = pd.read_excel(BytesIO(content), sheet_name="Data",
                             header=None, nrows=25, engine="openpyxl")
    for i, row in df_scan.iterrows():
        if "ID de contrato" in row.values:
            return i
    raise ValueError("No se encontró 'ID de contrato' en la hoja Data.")

@st.cache_data(show_spinner=False)
def cargar_pivot(fhash: str, content: bytes) -> pd.DataFrame:
    hi = detectar_header(content)
    df = pd.read_excel(BytesIO(content), sheet_name="Data", header=hi, engine="openpyxl")
    df = df.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)
    df = df[~df["Estado del contrato"].astype(str).str.lower().str.strip().isin(["cerrado","cerrados"])]
    df = df.rename(columns={k: v for k, v in PIVOT_COL_MAP.items() if k in df.columns})
    df["id"] = df["id"].astype(str).str.strip()
    for rw, pr in [("fecha_inicio_raw","fecha_inicio"),("fecha_termino_raw","fecha_termino")]:
        if rw in df.columns:
            df[pr] = df[rw].apply(parse_fecha)
    df["comprador_canon"] = df["propietario_raw"].apply(canon) if "propietario_raw" in df.columns else "Sin asignar"
    df["es_oficial"]      = df["comprador_canon"].apply(es_comprador_oficial)
    df["tipo_comprador"]  = df["comprador_canon"].apply(tipo_comprador)
    hoy = pd.Timestamp.today().normalize()
    df["dias_venc"] = (df["fecha_termino"] - hoy).dt.days if "fecha_termino" in df.columns else None
    def _indef(row):
        if norm(row.get("indefinido_raw","")) in ("sí","si","yes","1","true","indefinido"): return True
        ft = row.get("fecha_termino")
        return pd.notna(ft) and isinstance(ft, pd.Timestamp) and ft.year > 2100
    df["es_indefinido"] = df.apply(_indef, axis=1)
    df["riesgo"]        = df.apply(lambda r: calcular_riesgo(
        str(r.get("estado_ariba","")), r.get("dias_venc"), r.get("es_indefinido",False)), axis=1)
    df["tiene_garantia"] = df["garantia_ariba"].apply(
        lambda v: norm(v) in ("sí","si","yes")) if "garantia_ariba" in df.columns else False
    if "monto_total" in df.columns:
        df["monto_total"] = pd.to_numeric(df["monto_total"], errors="coerce").fillna(0)
    if "fecha_inicio" in df.columns:
        df["anio_inicio"] = df["fecha_inicio"].dt.year
    df["fuente"] = "Ariba"
    return df[df["id"].notna() & (df["id"] != "nan") & (df["id"] != "")].reset_index(drop=True)


@st.cache_data(show_spinner=False)
def cargar_consolidado(fhash: str, content: bytes) -> pd.DataFrame:
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    if "Consolidado de Contratos" not in wb.sheetnames:
        raise ValueError("No se encontró la hoja 'Consolidado de Contratos'.")
    ws   = wb["Consolidado de Contratos"]
    rows = list(ws.iter_rows(values_only=True))
    df   = pd.DataFrame(rows[1:], columns=rows[0]).dropna(how="all").reset_index(drop=True)
    df   = df.rename(columns={k: v for k, v in CONS_COL_MAP.items() if k in df.columns})
    df["id"] = df["id"].astype(str).str.strip() if "id" in df.columns else ""
    for col in ("fecha_termino_cons","venc_garantia"):
        if col in df.columns:
            df[col] = df[col].apply(parse_fecha)
    return df[df["id"].notna() & (df["id"] != "") & (df["id"] != "nan")].reset_index(drop=True)


# ──────────────────────────────────────────────────────────────
# UNIVERSO UNIFICADO
# ──────────────────────────────────────────────────────────────
def construir_universo(df_p: pd.DataFrame, df_c: pd.DataFrame | None) -> pd.DataFrame:
    if df_c is None:
        return df_p.copy()
    ids_ariba = set(df_p["id"].astype(str))
    ids_cons  = set(df_c["id"].astype(str))
    df_merged = df_p.merge(df_c, on="id", how="left", suffixes=("","_c"))
    df_merged.loc[df_merged["id"].isin(ids_cons), "fuente"] = "Ambos"
    ids_solo_cons = ids_cons - ids_ariba
    if ids_solo_cons:
        df_solo = df_c[df_c["id"].isin(ids_solo_cons)].copy()
        df_solo["fuente"]        = "Solo Consolidado"
        df_solo["estado_ariba"]  = df_solo.get("estado_cons_ariba", pd.Series(dtype=str))
        df_solo["proveedor"]     = df_solo.get("proveedor_cons",     pd.Series(dtype=str))
        df_solo["fecha_termino"] = df_solo.get("fecha_termino_cons", pd.Series(dtype="datetime64[ns]"))
        df_solo["tiene_garantia"] = df_solo.get("garantia_cons", pd.Series(dtype=str)).apply(
            lambda v: norm(v) in ("sí","si","yes","aplica","true"))
        df_solo["indefinido_raw"] = df_solo.get("indefinido_cons", pd.Series(dtype=str))
        df_solo["monto_total"]   = pd.to_numeric(
            df_solo.get("monto_garantia", pd.Series(dtype=float)), errors="coerce").fillna(0)
        hoy = pd.Timestamp.today().normalize()
        df_solo["dias_venc"] = (df_solo["fecha_termino"] - hoy).dt.days
        def _indef_cons(row):
            if norm(row.get("indefinido_raw","")) in ("sí","si","yes","1","true","indefinido","x"): return True
            ft = row.get("fecha_termino")
            return pd.notna(ft) and isinstance(ft, pd.Timestamp) and ft.year > 2100
        df_solo["es_indefinido"] = df_solo.apply(_indef_cons, axis=1)
        df_solo["riesgo"] = df_solo.apply(lambda r: calcular_riesgo(
            str(r.get("estado_ariba","")), r.get("dias_venc"), r.get("es_indefinido",False)), axis=1)
        def _comp_cons(row):
            ce = str(row.get("comprador_estrat","")).strip()
            ct = str(row.get("comprador_tact","")).strip()
            if ce and ce.lower() not in ("nan",""):  return canon(ce)
            if ct and ct.lower() not in ("nan",""):  return canon(ct)
            return "Sin asignar"
        df_solo["comprador_canon"] = df_solo.apply(_comp_cons, axis=1)
        df_solo["es_oficial"]      = df_solo["comprador_canon"].apply(es_comprador_oficial)
        df_solo["tipo_comprador"]  = df_solo["comprador_canon"].apply(tipo_comprador)
        df_solo["propietario_raw"] = df_solo["comprador_canon"]
        cols_comunes = [c for c in df_merged.columns if c in df_solo.columns]
        df_merged = pd.concat([df_merged, df_solo[cols_comunes]], ignore_index=True)
    return df_merged.reset_index(drop=True)


def comparar(df_p: pd.DataFrame, df_c: pd.DataFrame) -> pd.DataFrame:
    merged = df_p.merge(df_c, on="id", how="outer", suffixes=("","_c"), indicator=True)
    merged["es_nuevo_ariba"] = merged["_merge"] == "left_only"
    merged["es_solo_cons"]   = merged["_merge"] == "right_only"
    def _dif_comprador(r):
        if r["es_nuevo_ariba"] or r["es_solo_cons"]: return False
        comp_pivot = str(r.get("propietario_raw","")).strip().lower()
        comp_cons  = str(r.get("comprador_estrat", r.get("comprador_tact",""))).strip().lower()
        return comp_pivot and comp_cons and comp_pivot != comp_cons
    merged["dif_comprador"] = merged.apply(_dif_comprador, axis=1)
    merged["dif_estado"] = merged.apply(
        lambda r: not r["es_nuevo_ariba"] and not r["es_solo_cons"] and
                  norm(r.get("estado_ariba","")) != norm(r.get("estado_cons_ariba","")), axis=1)
    def _dif_fecha(r):
        if r["es_nuevo_ariba"] or r["es_solo_cons"]: return False
        fp, fc = r.get("fecha_termino"), r.get("fecha_termino_cons")
        if pd.isna(fp) or pd.isna(fc): return False
        if not (isinstance(fp, pd.Timestamp) and isinstance(fc, pd.Timestamp)): return False
        return abs((fp - fc).days) > 1
    merged["dif_fecha"] = merged.apply(_dif_fecha, axis=1)
    merged["dif_proveedor"] = merged.apply(
        lambda r: not r["es_nuevo_ariba"] and not r["es_solo_cons"] and
                  norm(r.get("proveedor","")) != norm(r.get("proveedor_cons","")), axis=1)
    merged["dif_garantia"] = merged.apply(
        lambda r: not r["es_nuevo_ariba"] and not r["es_solo_cons"] and
                  norm(r.get("garantia_ariba","")) != norm(r.get("garantia_cons","")), axis=1)
    def _status(r):
        if r["es_solo_cons"]:                        return "SOLO CONSOLIDADO"
        if r["es_nuevo_ariba"]:                      return "NUEVO EN ARIBA"
        if r["dif_estado"] or r["dif_fecha"]:        return "DESACTUALIZADO"
        if r["dif_proveedor"] or r["dif_garantia"]:  return "REVISAR"
        return "OK"
    merged["sync_status"] = merged.apply(_status, axis=1)
    def _cambios(r):
        if r["es_solo_cons"]:
            comp = r.get("comprador_estrat","") or r.get("comprador_tact","") or "—"
            return f"📂 Contrato registrado solo en el Consolidado (comprador: {comp}) — verificar si debe subirse a Ariba"
        if r["es_nuevo_ariba"]:
            return "🆕 Contrato nuevo en Ariba — no existe en el Consolidado"
        msgs = []
        if r["dif_comprador"]:
            msgs.append(f"👤 Comprador: Pivot=«{r.get('propietario_raw','—')}» / Consolidado=«{r.get('comprador_estrat', r.get('comprador_tact','—'))}»")
        if r["dif_estado"]:
            ea = r.get("estado_ariba","—"); ec = r.get("estado_cons_ariba","—") or "vacío"
            msgs.append(f"📄 Estado: Ariba=«{ea}» / Consolidado=«{ec}»")
        if r["dif_fecha"]:
            fp = r.get("fecha_termino"); fc = r.get("fecha_termino_cons")
            fps = fp.strftime("%d/%m/%Y") if pd.notna(fp) else "—"
            fcs = fc.strftime("%d/%m/%Y") if pd.notna(fc) else "—"
            msgs.append(f"📅 Fecha término: Ariba={fps} / Consolidado={fcs}")
        if r["dif_proveedor"]:
            msgs.append(f"🏢 Proveedor: «{r.get('proveedor','—')}» ≠ «{r.get('proveedor_cons','—')}»")
        if r["dif_garantia"]:
            msgs.append(f"🔒 Garantía: Ariba=«{r.get('garantia_ariba','—')}» / Consolidado=«{r.get('garantia_cons','—')}»")
        return " | ".join(msgs) if msgs else "✅ Sincronizado"
    merged["cambios"] = merged.apply(_cambios, axis=1)
    def _comp_merged(r):
        if r.get("es_solo_cons"):
            ce = str(r.get("comprador_estrat","")).strip()
            ct = str(r.get("comprador_tact","")).strip()
            if ce and ce.lower() not in ("nan",""):  return canon(ce)
            if ct and ct.lower() not in ("nan",""):  return canon(ct)
            return "Sin asignar"
        if r.get("dif_comprador"):
            ce = str(r.get("comprador_estrat","")).strip()
            ct = str(r.get("comprador_tact","")).strip()
            if ce and ce.lower() not in ("nan",""):  return canon(ce)
            if ct and ct.lower() not in ("nan",""):  return canon(ct)
        raw = r.get("propietario_raw","")
        return canon(raw) if raw else "Sin asignar"
    merged["comprador_canon"] = merged.apply(_comp_merged, axis=1)
    return merged


# ══════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("""
    <div style="padding:12px 4px 18px;border-bottom:1px solid rgba(255,255,255,0.2);margin-bottom:4px;">
      <div style="display:flex;align-items:center;gap:10px;">
        <div style="background:rgba(255,255,255,0.15);border-radius:10px;padding:7px 10px;
                    font-size:1.4rem;line-height:1;">📋</div>
        <div>
          <div style="font-size:1rem;font-weight:800;letter-spacing:-.01em;color:#ffffff;">
            Softys Chile
          </div>
          <div style="font-size:.67rem;color:rgba(232,242,251,0.7);margin-top:1px;font-weight:400;">
            Compras Estratégicas · Contratos
          </div>
        </div>
      </div>
    </div>
    <div style="height:8px"></div>
    """, unsafe_allow_html=True)

    # ── Leer estado persistente ────────────────────────────────
    piv_bytes_persist, piv_meta   = leer_pivot_persistente()
    cons_bytes_persist, cons_meta = leer_cons_persistente()
    od_cfg = _leer_od_config()

    # ══════════════════════════════════════════════════════════
    # SECCIÓN PIVOT
    # ══════════════════════════════════════════════════════════
    st.markdown(
        '<div style="font-size:.72rem;font-weight:700;text-transform:uppercase;'
        'letter-spacing:.08em;color:rgba(255,255,255,0.6);margin-bottom:6px;">'
        '📁 Pivot Ariba <span style="color:#7DD3FC;">(obligatorio)</span></div>',
        unsafe_allow_html=True)

    if piv_bytes_persist and piv_meta:
        ts_up  = piv_meta.get("uploaded_at","—")
        size_s = _fmt_size(piv_meta.get("size_bytes",0))
        fname  = piv_meta.get("filename","—")
        by_s   = piv_meta.get("uploaded_by","—")
        st.markdown(f"""
        <div style="background:rgba(0,166,81,0.15);border-radius:10px;border-left:3px solid #00A651;
                    padding:10px 12px;margin-bottom:8px;">
          <div style="font-size:.6rem;font-weight:700;text-transform:uppercase;letter-spacing:.08em;
                      color:#7DEFA3;margin-bottom:2px;">✅ Archivo guardado en servidor</div>
          <div style="font-size:.78rem;font-weight:700;color:#ffffff;word-break:break-all;">{fname}</div>
          <div style="font-size:.65rem;color:rgba(255,255,255,0.65);margin-top:3px;line-height:1.6;">
            📅 {ts_up}<br>💾 {size_s} &nbsp;·&nbsp; 🔑 MD5 OK<br>👤 sesión {by_s}
          </div>
        </div>
        """, unsafe_allow_html=True)

        with st.expander("🔄 Reemplazar Pivot", expanded=False):
            st.caption("Sube un nuevo archivo para sobrescribir el actual en el servidor.")
            up_pivot_new = st.file_uploader("Nuevo Pivot", type=["xlsx","xls"],
                                             key="piv_replace", label_visibility="collapsed")
            if up_pivot_new:
                up_pivot_new.seek(0)
                new_bytes = up_pivot_new.read()
                new_hash  = _md5(new_bytes)
                if new_hash != piv_meta.get("hash_md5",""):
                    if st.button("💾 Guardar nuevo Pivot", key="btn_save_piv"):
                        with st.spinner("Guardando…"):
                            guardar_pivot_persistente(new_bytes, up_pivot_new.name, _session_id())
                        st.success("✅ Pivot actualizado. Recargando…")
                        st.rerun()
                else:
                    st.info("ℹ️ El archivo subido es idéntico al guardado (mismo MD5).")

        col_del_piv, _ = st.columns([1,1])
        with col_del_piv:
            if st.button("🗑️ Eliminar Pivot", key="btn_del_piv",
                         help="Elimina el archivo del servidor"):
                eliminar_pivot_persistente()
                st.warning("Pivot eliminado del servidor.")
                st.rerun()

        piv_bytes = piv_bytes_persist
        up_pivot  = None

    else:
        up_pivot = st.file_uploader("Pivot", type=["xlsx","xls"],
                                     key="piv", label_visibility="collapsed")
        st.caption("Descarga directa de SAP Ariba Analysis")
        piv_bytes = None
        if up_pivot:
            up_pivot.seek(0)
            piv_bytes = up_pivot.read()
            if st.button("💾 Guardar en servidor", key="btn_save_piv_first"):
                with st.spinner("Guardando Pivot en servidor…"):
                    guardar_pivot_persistente(piv_bytes, up_pivot.name, _session_id())
                st.success("✅ Pivot guardado. La app cargará automáticamente desde ahora.")
                st.rerun()
            else:
                st.caption("⚡ El archivo se usa en esta sesión. Guárdalo para que persista entre sesiones.")

    # ══════════════════════════════════════════════════════════
    # SECCIÓN CONSOLIDADO — con integración OneDrive
    # ══════════════════════════════════════════════════════════
    st.markdown(
        '<div style="font-size:.72rem;font-weight:700;text-transform:uppercase;'
        'letter-spacing:.08em;color:rgba(255,255,255,0.6);margin-bottom:6px;margin-top:14px;">'
        '📄 Consolidado OneDrive <span style="color:rgba(255,255,255,0.5);">(opcional)</span></div>',
        unsafe_allow_html=True)

    # ── Tarjeta de estado si hay archivo persistente ──────────
    if cons_bytes_persist and cons_meta:
        ts_c    = cons_meta.get("uploaded_at","—")
        size_c  = _fmt_size(cons_meta.get("size_bytes",0))
        fname_c = cons_meta.get("filename","—")
        # Mostrar si tiene URL OneDrive configurada
        od_url_guardada = od_cfg.get("url","")
        od_ultimo_sync  = od_cfg.get("ultimo_sync_ok","—")
        if od_url_guardada:
            badge_od = f'<span style="font-size:.58rem;background:rgba(0,114,206,0.3);color:#7DD3FC;border-radius:4px;padding:1px 6px;margin-left:4px;">🔗 OneDrive vinculado</span>'
        else:
            badge_od = f'<span style="font-size:.58rem;background:rgba(255,255,255,0.1);color:rgba(255,255,255,0.5);border-radius:4px;padding:1px 6px;margin-left:4px;">📁 subida manual</span>'

        st.markdown(f"""
        <div style="background:rgba(108,63,196,0.18);border-radius:10px;border-left:3px solid #9F7AEA;
                    padding:10px 12px;margin-bottom:8px;">
          <div style="font-size:.6rem;font-weight:700;text-transform:uppercase;letter-spacing:.08em;
                      color:#C4B5FD;margin-bottom:2px;">✅ Consolidado guardado {badge_od}</div>
          <div style="font-size:.78rem;font-weight:700;color:#ffffff;word-break:break-all;">{fname_c}</div>
          <div style="font-size:.65rem;color:rgba(255,255,255,0.65);margin-top:3px;line-height:1.6;">
            📅 {ts_c} &nbsp;·&nbsp; 💾 {size_c}
            {"<br>🔄 Último sync OneDrive: " + od_ultimo_sync if od_url_guardada else ""}
          </div>
        </div>
        """, unsafe_allow_html=True)

        # ── Botón principal: Actualizar desde OneDrive ────────
        if od_url_guardada:
            if st.button("🔄 Actualizar desde OneDrive", key="btn_sync_od_main",
                         help="Descarga la última versión del archivo desde OneDrive y actualiza el dashboard"):
                with st.spinner("Conectando con OneDrive y descargando…"):
                    ok, msg = sincronizar_desde_onedrive(od_url_guardada, _session_id())
                if ok:
                    st.success(msg)
                    st.rerun()
                else:
                    st.error(msg)

        # ── Expander: Configurar / cambiar URL OneDrive ───────
        with st.expander("🔗 Configurar link OneDrive", expanded=not od_url_guardada):
            st.markdown(
                '<div style="font-size:.7rem;color:rgba(255,255,255,0.75);line-height:1.6;margin-bottom:8px;">'
                '📌 <strong>Cómo obtener el link:</strong><br>'
                '1. Abre el archivo en OneDrive / SharePoint<br>'
                '2. Clic en <strong>Compartir</strong><br>'
                '3. Selecciona <strong>"Cualquier persona con el link"</strong><br>'
                '4. Copia el link y pégalo abajo<br>'
                '<span style="color:#FCA5A5;">⚠️ Links de "Solo tu organización" requieren login y no funcionan aquí</span>'
                '</div>',
                unsafe_allow_html=True)

            url_input = st.text_input(
                "Link de OneDrive / SharePoint",
                value=od_url_guardada,
                placeholder="https://softys-my.sharepoint.com/:x:/r/...",
                key="od_url_input",
                label_visibility="collapsed")

            col_od1, col_od2 = st.columns([3, 2])
            with col_od1:
                if st.button("💾 Guardar link", key="btn_guardar_od_url"):
                    url_limpia = url_input.strip() if url_input else ""
                    if not url_limpia:
                        st.warning("⚠️ Escribe un link primero.")
                    elif "sharepoint.com" not in url_limpia and "onedrive.live.com" not in url_limpia and "1drv.ms" not in url_limpia:
                        st.error("❌ El link no parece ser de OneDrive o SharePoint.")
                    else:
                        guardar_url_onedrive(url_limpia)
                        st.success("✅ Link guardado.")
                        st.rerun()
            with col_od2:
                if od_url_guardada and st.button("🗑️ Quitar link", key="btn_del_od_url"):
                    eliminar_url_onedrive()
                    st.rerun()

            # Botón de prueba de conexión
            if url_input and url_input.strip():
                if st.button("🧪 Probar conexión", key="btn_test_od"):
                    with st.spinner("Probando descarga…"):
                        content_test, err_test = descargar_consolidado_onedrive(url_input.strip())
                    if content_test:
                        st.success(f"✅ Conexión OK — archivo descargado ({_fmt_size(len(content_test))})")
                    else:
                        st.error(err_test)

        # ── Expander: Reemplazar manualmente ──────────────────
        with st.expander("📁 Reemplazar con archivo local", expanded=False):
            st.caption("Alternativa: sube el archivo manualmente desde tu PC.")
            up_cons_new = st.file_uploader("Nuevo Consolidado", type=["xlsx","xls"],
                                            key="con_replace", label_visibility="collapsed")
            if up_cons_new:
                up_cons_new.seek(0)
                new_c_bytes = up_cons_new.read()
                if _md5(new_c_bytes) != cons_meta.get("hash_md5",""):
                    if st.button("💾 Guardar Consolidado", key="btn_save_cons"):
                        with st.spinner("Guardando…"):
                            guardar_cons_persistente(new_c_bytes, up_cons_new.name, _session_id())
                        st.success("✅ Consolidado actualizado.")
                        st.rerun()
                else:
                    st.info("ℹ️ Archivo idéntico al guardado.")

        col_del_cons, _ = st.columns([1,1])
        with col_del_cons:
            if st.button("🗑️ Eliminar Consolidado", key="btn_del_cons"):
                eliminar_cons_persistente()
                eliminar_url_onedrive()
                st.warning("Consolidado eliminado.")
                st.rerun()

        cons_bytes = cons_bytes_persist
        up_cons    = None

    else:
        # ── Sin archivo persistente: mostrar opciones de carga ─
        st.markdown(
            '<div style="font-size:.68rem;color:rgba(255,255,255,0.6);margin-bottom:8px;line-height:1.5;">'
            'Conecta el archivo desde OneDrive o súbelo manualmente.</div>',
            unsafe_allow_html=True)

        # ── Opción A: Link OneDrive ────────────────────────────
        with st.expander("🔗 Cargar desde OneDrive", expanded=True):
            st.markdown(
                '<div style="font-size:.7rem;color:rgba(255,255,255,0.75);line-height:1.6;margin-bottom:6px;">'
                '📌 El link debe ser <strong>"Cualquier persona con el link"</strong> desde SharePoint.<br>'
                '<span style="color:#FCA5A5;">⚠️ Links de "Solo tu organización" requieren login y no funcionan.</span>'
                '</div>',
                unsafe_allow_html=True)

            url_nueva = st.text_input(
                "Link OneDrive / SharePoint",
                placeholder="https://softys-my.sharepoint.com/:x:/r/...",
                key="od_url_nueva",
                label_visibility="collapsed")

            if url_nueva and url_nueva.strip():
                col_t1, col_t2 = st.columns(2)
                with col_t1:
                    if st.button("⬇️ Descargar y guardar", key="btn_od_descarga_nueva"):
                        url_n = url_nueva.strip()
                        if "sharepoint.com" not in url_n and "onedrive.live.com" not in url_n and "1drv.ms" not in url_n:
                            st.error("❌ El link no parece ser de OneDrive o SharePoint.")
                        else:
                            with st.spinner("Descargando desde OneDrive…"):
                                ok, msg = sincronizar_desde_onedrive(url_n, _session_id())
                            if ok:
                                guardar_url_onedrive(url_n)
                                st.success(msg + "\n\nLink guardado para futuras actualizaciones.")
                                st.rerun()
                            else:
                                st.error(msg)
                with col_t2:
                    if st.button("🧪 Probar link", key="btn_test_od_nueva"):
                        with st.spinner("Probando…"):
                            c_test, e_test = descargar_consolidado_onedrive(url_nueva.strip())
                        if c_test:
                            st.success(f"✅ OK ({_fmt_size(len(c_test))})")
                        else:
                            st.error(e_test)

        # ── Opción B: Subida manual ────────────────────────────
        with st.expander("📁 Subir archivo manualmente", expanded=False):
            up_cons = st.file_uploader("Consolidado", type=["xlsx","xls"],
                                        key="con", label_visibility="collapsed")
            st.caption("Descarga el archivo de SharePoint y súbelo aquí.")
            cons_bytes = None
            if up_cons:
                up_cons.seek(0)
                cons_bytes = up_cons.read()
                if st.button("💾 Guardar Consolidado en servidor", key="btn_save_cons_first"):
                    with st.spinner("Guardando Consolidado…"):
                        guardar_cons_persistente(cons_bytes, up_cons.name, _session_id())
                    st.success("✅ Consolidado guardado.")
                    st.rerun()
                else:
                    st.caption("⚡ Usado solo en esta sesión. Guárdalo para persistir.")
        up_cons = None  # evitar referencia suelta

    st.markdown("---")
    filtros_ph = st.empty()

# ══════════════════════════════════════════════════════════════
# PANTALLA DE BIENVENIDA
# ══════════════════════════════════════════════════════════════
_tiene_pivot = (piv_bytes is not None) or (piv_bytes_persist is not None)
_bytes_pivot  = piv_bytes if piv_bytes is not None else piv_bytes_persist

if not _tiene_pivot or _bytes_pivot is None:
    st.markdown("""
    <div style="display:flex;flex-direction:column;align-items:center;
         justify-content:center;padding:70px 30px 40px;text-align:center;">
      <div style="background:linear-gradient(135deg,#005CA9,#0072CE);
                  border-radius:16px;padding:16px 24px;margin-bottom:24px;
                  box-shadow:0 8px 24px rgba(0,92,169,.25);">
        <span style="font-size:2rem;">📋</span>
        <span style="color:#ffffff;font-size:1rem;font-weight:800;margin-left:10px;
                     letter-spacing:-.01em;vertical-align:middle;">
          Softys · Gestión de Contratos
        </span>
      </div>
      <h1 style="font-size:1.55rem;font-weight:800;color:#1A2E44;margin:0 0 10px;line-height:1.25;">
        Dashboard de Compras Estratégicas
      </h1>
      <p style="color:#5A7490;max-width:500px;line-height:1.7;margin-bottom:30px;font-size:.93rem;">
        Sube el <strong style="color:#005CA9;">Pivot de Ariba</strong> para ver indicadores.<br>
        Agrega el <strong style="color:#6C3FC4;">Consolidado del Drive</strong> para detectar
        qué está desactualizado y quién debe actualizarlo.
      </p>
      <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:14px;max-width:680px;width:100%;">
        <div style="background:#ffffff;border-radius:12px;padding:16px 12px;
                    border-top:3px solid #00A651;box-shadow:0 2px 8px rgba(0,0,0,.06);">
          <div style="font-size:1.5rem;">📊</div>
          <div style="font-weight:700;font-size:.8rem;color:#1A2E44;margin-top:7px;">KPIs automáticos</div>
          <div style="font-size:.72rem;color:#8FA3B8;margin-top:3px;">10 indicadores clave</div>
        </div>
        <div style="background:#ffffff;border-radius:12px;padding:16px 12px;
                    border-top:3px solid #0072CE;box-shadow:0 2px 8px rgba(0,0,0,.06);">
          <div style="font-size:1.5rem;">🔄</div>
          <div style="font-weight:700;font-size:.8rem;color:#1A2E44;margin-top:7px;">Sincronización</div>
          <div style="font-size:.72rem;color:#8FA3B8;margin-top:3px;">Campo a campo</div>
        </div>
        <div style="background:#ffffff;border-radius:12px;padding:16px 12px;
                    border-top:3px solid #6C3FC4;box-shadow:0 2px 8px rgba(0,0,0,.06);">
          <div style="font-size:1.5rem;">📂</div>
          <div style="font-weight:700;font-size:.8rem;color:#1A2E44;margin-top:7px;">Contratos Drive</div>
          <div style="font-size:.72rem;color:#8FA3B8;margin-top:3px;">Visibles sin Ariba</div>
        </div>
        <div style="background:#ffffff;border-radius:12px;padding:16px 12px;
                    border-top:3px solid #F59E0B;box-shadow:0 2px 8px rgba(0,0,0,.06);">
          <div style="font-size:1.5rem;">👤</div>
          <div style="font-weight:700;font-size:.8rem;color:#1A2E44;margin-top:7px;">Por comprador</div>
          <div style="font-size:.72rem;color:#8FA3B8;margin-top:3px;">Alertas personalizadas</div>
        </div>
      </div>
      <div style="margin-top:32px;padding:12px 20px;background:#EAF2FB;
                  border-radius:99px;font-size:.78rem;color:#005CA9;font-weight:600;">
        ← Sube el Pivot de Ariba en la barra lateral para comenzar
      </div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ──────────────────────────────────────────────────────────────
# CARGA DE DATOS
# ──────────────────────────────────────────────────────────────
with st.spinner("Procesando Pivot de Ariba..."):
    try:
        df_piv = cargar_pivot(_md5(_bytes_pivot), _bytes_pivot)
    except Exception as e:
        st.error(f"❌ Error al leer el Pivot: {e}"); st.stop()

df_cons_raw = None
_bytes_cons = cons_bytes if cons_bytes is not None else cons_bytes_persist
if _bytes_cons:
    with st.spinner("Cargando Consolidado..."):
        try:
            df_cons_raw = cargar_consolidado(_md5(_bytes_cons), _bytes_cons)
        except Exception as e:
            st.warning(f"⚠️ No se pudo leer el Consolidado: {e}")

with st.spinner("Unificando fuentes..."):
    df_universo = construir_universo(df_piv, df_cons_raw)

# ──────────────────────────────────────────────────────────────
# FILTROS
# ──────────────────────────────────────────────────────────────
_piv_name  = (piv_meta.get("filename","pivot.xlsx") if piv_meta
              else (up_pivot.name if up_pivot else "pivot.xlsx"))
_cons_name = (cons_meta.get("filename","consolidado.xlsx") if cons_meta else "consolidado.xlsx")

with st.sidebar:
    with filtros_ph.container():
        st.markdown(
            '<div style="font-size:.72rem;font-weight:700;text-transform:uppercase;'
            'letter-spacing:.08em;color:rgba(255,255,255,0.6);margin-bottom:8px;">🎛️ Filtros</div>',
            unsafe_allow_html=True)

        mostrar_solo_oficiales = st.checkbox("Solo compradores oficiales", value=True,
            help="Filtra contratos cuyo propietario es un comprador registrado")

        f_fuente = st.selectbox("📂 Fuente", ["Todas","Ariba","Ambos","Solo Consolidado"])
        f_riesgo  = st.selectbox("🚦 Riesgo", ["Todos"] + sorted(df_universo["riesgo"].dropna().unique()))
        f_tipo    = st.selectbox("👥 Tipo comprador",
                                  ["Todos","Estratégico","Táctico","Estratégico + Táctico","No registrado"])
        compradores_lista = sorted(df_universo["comprador_canon"].dropna().unique().tolist())
        f_comp    = st.selectbox("👤 Comprador", ["Todos"] + compradores_lista)
        f_estado  = st.selectbox("📄 Estado Ariba",
                                  ["Todos"] + sorted(df_universo["estado_ariba"].dropna().unique()))
        f_gar     = st.selectbox("🔒 Garantía", ["Todas","Con garantía","Sin garantía"])
        f_indef   = st.selectbox("♾️ Indefinidos", ["Todos","Solo indefinidos","Solo con fecha"])

        st.markdown("---")
        n_solo_cons = (df_universo["fuente"] == "Solo Consolidado").sum()
        piv_ts_label = f"\n📅 {piv_meta['uploaded_at']}" if piv_meta else ""
        st.caption(f"📁 {_piv_name}{piv_ts_label}\n{len(df_piv):,} contratos activos en Ariba")
        if df_cons_raw is not None:
            od_url_sidebar = od_cfg.get("url","")
            od_sync_label  = f"\n🔗 OneDrive · último sync: {od_cfg.get('ultimo_sync_ok','—')}" if od_url_sidebar else ""
            st.caption(f"📄 {_cons_name}{od_sync_label}\n{len(df_cons_raw):,} filas · {n_solo_cons:,} solo en Drive")

# Aplicar filtros
df = df_universo.copy()
if mostrar_solo_oficiales:    df = df[df["es_oficial"]]
if f_fuente != "Todas":       df = df[df["fuente"] == f_fuente]
if f_riesgo != "Todos":       df = df[df["riesgo"] == f_riesgo]
if f_tipo   != "Todos":       df = df[df["tipo_comprador"] == f_tipo]
if f_comp   != "Todos":       df = df[df["comprador_canon"] == f_comp]
if f_estado != "Todos":       df = df[df["estado_ariba"] == f_estado]
if f_gar == "Con garantía":   df = df[df["tiene_garantia"]]
elif f_gar == "Sin garantía": df = df[~df["tiene_garantia"]]
if f_indef == "Solo indefinidos": df = df[df["es_indefinido"]]
elif f_indef == "Solo con fecha": df = df[~df["es_indefinido"]]

if df.empty:
    st.warning("⚠️ Ningún contrato coincide con los filtros."); st.stop()

# ──────────────────────────────────────────────────────────────
# ENCABEZADO
# ──────────────────────────────────────────────────────────────
n_solo_cons_vis = (df["fuente"] == "Solo Consolidado").sum()
pills = []
if df_cons_raw is not None:
    od_url_h = od_cfg.get("url","")
    if od_url_h:
        pills.append(
            '<span style="background:#EAF2FB;color:#005CA9;border-radius:99px;padding:3px 11px;'
            'font-size:.68rem;font-weight:700;margin-left:8px;border:1px solid #B8D4EF;">'
            '🔗 OneDrive vinculado</span>')
    pills.append(
        '<span style="background:#EAF2FB;color:#005CA9;border-radius:99px;padding:3px 11px;'
        'font-size:.68rem;font-weight:700;margin-left:5px;border:1px solid #B8D4EF;">'
        '🔄 Sincronización activa</span>')
if n_solo_cons_vis > 0:
    pills.append(
        f'<span style="background:#F0EAFF;color:#6C3FC4;border-radius:99px;padding:3px 11px;'
        f'font-size:.68rem;font-weight:700;margin-left:5px;border:1px solid #D4C5F0;">'
        f'📂 {n_solo_cons_vis} solo en Drive</span>')
pills.append(
    '<span style="background:#F0FDF4;color:#00703A;border-radius:99px;padding:3px 11px;'
    'font-size:.68rem;font-weight:700;margin-left:5px;border:1px solid #A7F3D0;">'
    '💾 Datos persistentes</span>')

st.markdown(f"""
<div style="display:flex;justify-content:space-between;align-items:flex-end;
            margin-bottom:16px;padding-bottom:14px;border-bottom:2px solid #D1DCE8;">
  <div>
    <div style="font-size:.68rem;font-weight:700;text-transform:uppercase;letter-spacing:.1em;
                color:#0072CE;margin-bottom:4px;">Compras Estratégicas · Chile</div>
    <h1 style="font-size:1.35rem;font-weight:800;color:#1A2E44;margin:0;line-height:1.2;">
      Gestión de Contratos{''.join(pills)}
    </h1>
    <div style="color:#7A96AF;font-size:.75rem;margin-top:5px;font-weight:400;">
      Fuentes: SAP Ariba + Consolidado Drive &nbsp;·&nbsp;
      <strong style="color:#1A2E44;">{df["id"].nunique():,}</strong> contratos en vista
    </div>
  </div>
  <div style="text-align:right;">
    <div style="font-size:.68rem;color:#A0B4C4;">
      {datetime.now().strftime('%d/%m/%Y')}<br>
      <span style="font-size:.62rem;">{datetime.now().strftime('%H:%M')}</span>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────
# TABS
# ──────────────────────────────────────────────────────────────
if df_cons_raw is not None:
    tab_kpi, tab_sync, tab_comp, tab_prov, tab_gar_tab, tab_exp = st.tabs([
        "📊 Resumen & KPIs", "🔄 Sincronización", "📌 Por Comprador",
        "🏢 Proveedores", "🔒 Garantías", "🔍 Explorador"])
else:
    tab_kpi, tab_comp, tab_prov, tab_gar_tab, tab_exp = st.tabs([
        "📊 Resumen & KPIs", "📌 Por Comprador",
        "🏢 Proveedores", "🔒 Garantías", "🔍 Explorador"])
    tab_sync = None

def _layout(fig, title="", h=275):
    fig.update_layout(
        title=dict(text=title, font=dict(size=12, color="#1A2E44", family="Inter"), x=0.01),
        paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
        font=dict(family="Inter", size=10, color="#2D3748"),
        margin=dict(t=36 if title else 12, b=10, l=8, r=8), height=h)
    return fig

# ══════════════════════════════════════════════
# TAB: KPIs & RESUMEN
# ══════════════════════════════════════════════
with tab_kpi:
    total   = len(df)
    bajo    = (df["riesgo"] == "BAJO 🟢").sum()
    medio   = (df["riesgo"] == "MEDIO 🟡").sum()
    alto    = (df["riesgo"] == "ALTO 🔴").sum()
    revisar = (df["riesgo"] == "REVISAR ⚪").sum()
    indef   = df["es_indefinido"].sum()
    gar     = df["tiene_garantia"].sum()
    monto   = df["monto_total"].sum() if "monto_total" in df.columns else 0
    pct_v   = f"{bajo/total*100:.0f}%" if total else "—"
    n_sc    = (df["fuente"] == "Solo Consolidado").sum()

    st.markdown(f"""
    <div class="kpi-row">
      <div class="kpi"><div class="kpi-lbl">📋 Total contratos</div>
        <div class="kpi-val">{total:,}</div><div class="kpi-sub">Ariba + Drive unificados</div></div>
      <div class="kpi g"><div class="kpi-lbl">✅ Vigentes</div>
        <div class="kpi-val">{bajo:,}</div><div class="kpi-sub">{pct_v} del total</div></div>
      <div class="kpi y"><div class="kpi-lbl">⚠️ Riesgo medio</div>
        <div class="kpi-val">{medio:,}</div><div class="kpi-sub">vencen ≤ 60 días</div></div>
      <div class="kpi r"><div class="kpi-lbl">🚨 Riesgo alto</div>
        <div class="kpi-val">{alto:,}</div><div class="kpi-sub">vencidos / cancelados</div></div>
      <div class="kpi gr"><div class="kpi-lbl">🔍 Por revisar</div>
        <div class="kpi-val">{revisar:,}</div><div class="kpi-sub">borrador / sin fecha</div></div>
    </div>
    <div class="kpi-row">
      <div class="kpi p"><div class="kpi-lbl">📂 Solo en Drive</div>
        <div class="kpi-val">{n_sc:,}</div><div class="kpi-sub">no están en Ariba</div></div>
      <div class="kpi"><div class="kpi-lbl">♾️ Indefinidos</div>
        <div class="kpi-val">{indef:,}</div><div class="kpi-sub">sin fecha de término</div></div>
      <div class="kpi g"><div class="kpi-lbl">🔒 Con garantía</div>
        <div class="kpi-val">{gar:,}</div><div class="kpi-sub">aplica boleta</div></div>
      <div class="kpi"><div class="kpi-lbl">👤 Compradores</div>
        <div class="kpi-val">{df["comprador_canon"].nunique():,}</div><div class="kpi-sub">propietarios únicos</div></div>
      <div class="kpi"><div class="kpi-lbl">💰 Monto total</div>
        <div class="kpi-val">{fmt_m(monto)}</div><div class="kpi-sub">CLP contratos filtrados</div></div>
    </div>
    """, unsafe_allow_html=True)

    if n_sc > 0:
        st.markdown(f"""
        <div class="alert-card purple">
          <strong>📂 {n_sc} contratos registrados solo en el Consolidado del Drive</strong><br>
          Son gestionados directamente por los compradores y <em>no tienen correspondencia en Ariba</em>.
          Están incluidos en todos los KPIs y filtros.
        </div>
        """, unsafe_allow_html=True)

    c1, c2, c3 = st.columns([1, 1.4, 1.2])
    with c1:
        d = df["riesgo"].value_counts().reset_index(); d.columns = ["r","n"]
        fig = go.Figure(go.Pie(
            labels=d["r"], values=d["n"], hole=0.58,
            marker_colors=[RIESGO_COLORES.get(r,"#999") for r in d["r"]],
            marker_line=dict(color="#ffffff", width=2),
            textinfo="percent+value", textfont=dict(size=10, family="Inter"),
            hovertemplate="<b>%{label}</b><br>%{value} contratos (%{percent})<extra></extra>"))
        fig.update_layout(
            annotations=[dict(text=f"<b>{total}</b>", x=0.5, y=0.5, font_size=18,
                              font_color="#1A2E44", font_family="Inter", showarrow=False)],
            legend=dict(font=dict(size=9, family="Inter"), orientation="h", y=-0.15),
            paper_bgcolor="#ffffff", margin=dict(t=36,b=44,l=8,r=8), height=280,
            title=dict(text="Distribución de riesgo",
                       font=dict(size=12,color="#1A2E44",family="Inter"), x=0.02))
        st.plotly_chart(fig, use_container_width=True)

    with c2:
        if "fecha_termino" in df.columns:
            hoy = pd.Timestamp.today().normalize()
            df_tl = df[df["fecha_termino"].notna() & df["dias_venc"].between(-30,180)].copy()
            if not df_tl.empty:
                df_tl["mes"] = df_tl["fecha_termino"].dt.to_period("M").astype(str)
                agr = df_tl.groupby(["mes","riesgo"]).size().reset_index(name="n")
                fig2 = px.bar(agr, x="mes", y="n", color="riesgo",
                    color_discrete_map=RIESGO_COLORES, barmode="stack",
                    labels={"mes":"","n":"Contratos","riesgo":"Riesgo"})
                fig2.update_traces(marker_line_width=0)
                fig2.update_layout(
                    xaxis=dict(tickangle=-30, gridcolor="#F0F4F8", tickfont=dict(size=9)),
                    yaxis=dict(gridcolor="#F0F4F8", tickfont=dict(size=9)),
                    legend=dict(orientation="h", y=-0.28, font=dict(size=9, family="Inter")),
                    paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
                    font=dict(family="Inter", size=10),
                    margin=dict(t=36,b=65,l=8,r=8), height=280,
                    title=dict(text="Vencimientos próximos 6 meses",
                               font=dict(size=12,color="#1A2E44",family="Inter"), x=0.02))
                st.plotly_chart(fig2, use_container_width=True)

    with c3:
        d_fuente = df["fuente"].value_counts().reset_index(); d_fuente.columns = ["f","n"]
        col_fuente = {"Ariba":"#0072CE", "Ambos":"#00A651", "Solo Consolidado":"#6C3FC4"}
        fig_f = go.Figure(go.Pie(
            labels=d_fuente["f"], values=d_fuente["n"], hole=0.58,
            marker_colors=[col_fuente.get(f,"#999") for f in d_fuente["f"]],
            marker_line=dict(color="#ffffff", width=2),
            textinfo="percent+value", textfont=dict(size=10, family="Inter")))
        fig_f.update_layout(
            legend=dict(font=dict(size=9, family="Inter"), orientation="h", y=-0.15),
            paper_bgcolor="#ffffff", margin=dict(t=36,b=44,l=8,r=8), height=280,
            title=dict(text="Origen de contratos",
                       font=dict(size=12,color="#1A2E44",family="Inter"), x=0.02))
        st.plotly_chart(fig_f, use_container_width=True)

    st.markdown('<div class="sec">🚨 Contratos que requieren acción inmediata</div>',
                unsafe_allow_html=True)
    df_alt = df[df["riesgo"].isin(["ALTO 🔴","MEDIO 🟡"])].copy()
    if not df_alt.empty:
        cols_a = [c for c in ["id","proveedor","comprador_canon","estado_ariba","dias_venc",
                               "riesgo","tiene_garantia","fuente"] if c in df_alt]
        ren    = {"id":"Contrato","proveedor":"Proveedor","comprador_canon":"Comprador",
                  "estado_ariba":"Estado","dias_venc":"Días","riesgo":"Riesgo",
                  "tiene_garantia":"Garantía","fuente":"Fuente"}
        tbl = df_alt[cols_a].rename(columns=ren).sort_values("Días")
        def hl(v):
            if "ALTO"  in str(v): return "background:#FEF2F2;color:#7F1D1D;font-weight:600"
            if "MEDIO" in str(v): return "background:#FFFBEB;color:#78350F;font-weight:600"
            if str(v) == "Solo Consolidado": return "background:#F5F0FF;color:#3B1A78;font-weight:600"
            return ""
        st.dataframe(
            tbl.style.map(hl, subset=[c for c in ["Riesgo","Fuente"] if c in tbl.columns])
               .format({"Días":"{:.0f}"}, na_rep="—"),
            use_container_width=True, height=240)
        a1,a2 = st.columns(2)
        a1.error(f"🚨 **{(df['riesgo']=='ALTO 🔴').sum()}** contratos en riesgo ALTO — acción urgente")
        a2.warning(f"⚠️ **{(df['riesgo']=='MEDIO 🟡').sum()}** contratos próximos a vencer")
    else:
        st.success("✅ Sin contratos críticos con los filtros actuales.")


# ══════════════════════════════════════════════
# TAB: SINCRONIZACIÓN
# ══════════════════════════════════════════════
if tab_sync is not None:
    with tab_sync:
        st.markdown("""
        <div class="alert-card blue">
          <strong>¿Cómo funciona?</strong><br>
          Compara <strong>ambas fuentes en igualdad</strong>: detecta contratos nuevos en Ariba
          (deben agregarse al Consolidado), contratos solo en el Drive (gestionados por compradores,
          no subidos a Ariba), y contratos en ambos con datos distintos.
        </div>
        """, unsafe_allow_html=True)

        with st.spinner("Comparando archivos..."):
            df_cmp = comparar(df_piv, df_cons_raw)

        n_ok      = (df_cmp["sync_status"] == "OK").sum()
        n_desact  = (df_cmp["sync_status"] == "DESACTUALIZADO").sum()
        n_nuevo_a = (df_cmp["sync_status"] == "NUEVO EN ARIBA").sum()
        n_solo_c  = (df_cmp["sync_status"] == "SOLO CONSOLIDADO").sum()
        n_rev     = (df_cmp["sync_status"] == "REVISAR").sum()
        total_c   = len(df_cmp)
        pct_ok    = f"{n_ok/total_c*100:.0f}%" if total_c else "—"

        st.markdown(f"""
        <div class="kpi-row">
          <div class="kpi g"><div class="kpi-lbl">✅ Sincronizados</div>
            <div class="kpi-val">{n_ok:,}</div><div class="kpi-sub">{pct_ok} del universo</div></div>
          <div class="kpi r"><div class="kpi-lbl">⚠️ Desactualizados</div>
            <div class="kpi-val">{n_desact:,}</div><div class="kpi-sub">estado o fecha difieren</div></div>
          <div class="kpi b"><div class="kpi-lbl">🆕 Nuevos en Ariba</div>
            <div class="kpi-val">{n_nuevo_a:,}</div><div class="kpi-sub">faltan en el Drive</div></div>
          <div class="kpi p"><div class="kpi-lbl">📂 Solo en Drive</div>
            <div class="kpi-val">{n_solo_c:,}</div><div class="kpi-sub">no están en Ariba</div></div>
          <div class="kpi y"><div class="kpi-lbl">🔍 Revisar</div>
            <div class="kpi-val">{n_rev:,}</div><div class="kpi-sub">proveedor / garantía</div></div>
        </div>
        """, unsafe_allow_html=True)

        if n_solo_c > 0:
            st.markdown(f"""
            <div class="alert-card purple">
              <strong>📂 {n_solo_c} contratos registrados solo en el Consolidado del Drive</strong><br>
              Se muestran en la lista de cada comprador y en el Explorador.
            </div>
            """, unsafe_allow_html=True)

        st.markdown('<div class="sec">🔔 Alertas por comprador</div>', unsafe_allow_html=True)
        problemas = df_cmp[df_cmp["sync_status"] != "OK"].copy()
        compradores_con_prob = sorted(problemas["comprador_canon"].dropna().unique())

        if not compradores_con_prob:
            st.success("🎉 ¡El Consolidado está completamente sincronizado con el Pivot de Ariba!")
        else:
            f_sync_comp = st.selectbox(
                "Ver alertas de:",
                ["Todos los compradores"] + compradores_con_prob,
                key="f_sync_comp")
            df_prob_view = (problemas if f_sync_comp == "Todos los compradores"
                            else problemas[problemas["comprador_canon"] == f_sync_comp])

            for comp in (compradores_con_prob if f_sync_comp == "Todos los compradores"
                         else [f_sync_comp]):
                grp = df_prob_view[df_prob_view["comprador_canon"] == comp]
                if grp.empty: continue
                n_grp = len(grp)
                n_d   = (grp["sync_status"] == "DESACTUALIZADO").sum()
                n_n   = (grp["sync_status"] == "NUEVO EN ARIBA").sum()
                n_sc2 = (grp["sync_status"] == "SOLO CONSOLIDADO").sum()
                n_r   = (grp["sync_status"] == "REVISAR").sum()
                tipo  = tipo_comprador(comp)
                es_of = es_comprador_oficial(comp)
                badge_tipo = (
                    f"<span style='background:#EAF2FB;color:#005CA9;border-radius:6px;"
                    f"padding:1px 8px;font-size:.67rem;margin-left:6px;font-weight:700;'>{tipo}</span>"
                    if es_of else
                    "<span style='background:#FFFBEB;color:#78350F;border-radius:6px;"
                    "padding:1px 8px;font-size:.67rem;margin-left:6px;font-weight:700;'>No registrado</span>")
                severity = "red" if n_d > 0 or n_n > 0 else ("purple" if n_sc2 > 0 else "yellow")
                partes = []
                if n_d:   partes.append(f"<strong>{n_d}</strong> desactualizados")
                if n_n:   partes.append(f"<strong>{n_n}</strong> nuevos en Ariba sin registrar en Drive")
                if n_sc2: partes.append(f"<strong>{n_sc2}</strong> gestionados solo en Drive")
                if n_r:   partes.append(f"<strong>{n_r}</strong> por revisar")
                resumen_html = " &nbsp;·&nbsp; ".join(partes)

                with st.expander(
                        f"👤 {comp}{badge_tipo}  —  {n_grp} contrato(s) con diferencias",
                        expanded=(n_d+n_n > 0)):
                    st.markdown(
                        f'<div class="alert-card {severity}"><strong>Situación para {comp}:'
                        f'</strong><br>{resumen_html}</div>',
                        unsafe_allow_html=True)
                    cols_det = [c for c in [
                        "id","proveedor","proveedor_cons","estado_ariba","estado_cons_ariba",
                        "fecha_termino","fecha_termino_cons","sync_status","cambios"]
                        if c in grp.columns]
                    ren_det  = {
                        "id":"Contrato","proveedor":"Proveedor (Ariba)",
                        "proveedor_cons":"Proveedor (Drive)",
                        "estado_ariba":"Estado Ariba","estado_cons_ariba":"Estado en Drive",
                        "fecha_termino":"Fecha Ariba","fecha_termino_cons":"Fecha en Drive",
                        "sync_status":"Estado Sync","cambios":"Detalle"}
                    tbl_det = grp[cols_det].rename(columns=ren_det)
                    def hl_sync_row(val):
                        bg = SYNC_BG.get(str(val),""); fg = SYNC_FG.get(str(val),"")
                        return f"background:{bg};color:{fg};font-weight:600" if bg else ""
                    st.dataframe(
                        tbl_det.style.map(
                            hl_sync_row,
                            subset=["Estado Sync"] if "Estado Sync" in tbl_det.columns else []),
                        use_container_width=True,
                        height=min(300, 60 + len(grp)*38))

        st.markdown('<div class="sec">📊 Estado de sincronización por comprador</div>',
                    unsafe_allow_html=True)
        sc = df_cmp.groupby(["comprador_canon","sync_status"]).size().reset_index(name="n")
        fig_sc = px.bar(sc, y="comprador_canon", x="n", color="sync_status",
            color_discrete_map=SYNC_COLORES, barmode="stack", orientation="h",
            labels={"comprador_canon":"","n":"Contratos","sync_status":"Estado"})
        fig_sc.update_traces(marker_line_width=0)
        fig_sc.update_layout(
            xaxis=dict(gridcolor="#F0F4F8", tickfont=dict(size=9)),
            yaxis=dict(categoryorder="total ascending", tickfont=dict(size=9)),
            legend=dict(orientation="h", y=-0.12, font=dict(size=9, family="Inter")),
            paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
            font=dict(family="Inter", size=10),
            height=max(280, df_cmp["comprador_canon"].nunique()*24),
            margin=dict(t=10,b=60,l=10,r=10))
        st.plotly_chart(fig_sc, use_container_width=True)

        st.markdown('<div class="sec">📥 Exportar reporte de sincronización</div>',
                    unsafe_allow_html=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        cols_exp = [c for c in [
            "id","proveedor","proveedor_cons","comprador_canon",
            "estado_ariba","estado_cons_ariba",
            "fecha_termino","fecha_termino_cons","sync_status","cambios"]
            if c in df_cmp.columns]
        ren_exp  = {
            "id":"Contrato","proveedor":"Proveedor Ariba","proveedor_cons":"Proveedor Drive",
            "comprador_canon":"Comprador","estado_ariba":"Estado Ariba",
            "estado_cons_ariba":"Estado Drive","fecha_termino":"Fecha Ariba",
            "fecha_termino_cons":"Fecha Drive","sync_status":"Estado Sync","cambios":"Detalle"}
        e1,e2,e3,e4 = st.columns(4)
        with e1:
            pend = df_cmp[df_cmp["sync_status"] != "OK"][cols_exp].rename(columns=ren_exp)
            st.download_button("⚠️ Todos los pendientes",
                               pend.to_csv(index=False).encode("utf-8-sig"),
                               f"sync_pendientes_{ts}.csv", mime="text/csv")
        with e2:
            nuev = df_cmp[df_cmp["sync_status"]=="NUEVO EN ARIBA"][cols_exp].rename(columns=ren_exp)
            st.download_button("🆕 Nuevos en Ariba",
                               nuev.to_csv(index=False).encode("utf-8-sig"),
                               f"sync_nuevos_ariba_{ts}.csv", mime="text/csv")
        with e3:
            solo_c_exp = df_cmp[df_cmp["sync_status"]=="SOLO CONSOLIDADO"][cols_exp].rename(columns=ren_exp)
            st.download_button("📂 Solo en Drive",
                               solo_c_exp.to_csv(index=False).encode("utf-8-sig"),
                               f"sync_solo_drive_{ts}.csv", mime="text/csv")
        with e4:
            desact = df_cmp[df_cmp["sync_status"]=="DESACTUALIZADO"][cols_exp].rename(columns=ren_exp)
            st.download_button("🔄 Estado/fecha diferente",
                               desact.to_csv(index=False).encode("utf-8-sig"),
                               f"sync_desact_{ts}.csv", mime="text/csv")


# ══════════════════════════════════════════════
# TAB: POR COMPRADOR
# ══════════════════════════════════════════════
with tab_comp:
    st.markdown("""
    <div class="alert-card blue" style="margin-bottom:14px;">
      Los contratos se muestran desde <strong>ambas fuentes</strong>: Ariba y Consolidado del Drive.
      Los contratos gestionados solo en el Drive aparecen con badge
      <span class="badge-cons">📂 Solo Drive</span>.
    </div>
    """, unsafe_allow_html=True)

    dc = df.groupby(["comprador_canon","riesgo"]).size().reset_index(name="n")
    orden = df["comprador_canon"].value_counts().index.tolist()
    dc["comprador_canon"] = pd.Categorical(dc["comprador_canon"],
                                            categories=orden[::-1], ordered=True)
    dc = dc.sort_values("comprador_canon")
    fig_c = px.bar(dc, y="comprador_canon", x="n", color="riesgo",
        color_discrete_map=SYNC_COLORES, barmode="stack", orientation="h",
        labels={"comprador_canon":"","n":"Contratos","riesgo":"Riesgo"})
    fig_c.update_traces(marker_line_width=0)
    fig_c.update_layout(
        xaxis=dict(gridcolor="#F0F4F8", tickfont=dict(size=9)),
        yaxis=dict(tickfont=dict(size=9)),
        legend=dict(orientation="h", y=-0.12, font=dict(size=9, family="Inter")),
        paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
        font=dict(family="Inter", size=10),
        title=dict(text="Contratos por comprador (Ariba + Drive)",
                   font=dict(size=12,color="#1A2E44",family="Inter"), x=0.01),
        height=max(260, df["comprador_canon"].nunique()*26),
        margin=dict(t=38,b=60,l=10,r=10))
    st.plotly_chart(fig_c, use_container_width=True)

    resumen = df.groupby(["comprador_canon","tipo_comprador"]).agg(
        Contratos    =("id","count"),
        Solo_Drive   =("fuente", lambda x: (x=="Solo Consolidado").sum()),
        Riesgo_Alto  =("riesgo", lambda x: (x=="ALTO 🔴").sum()),
        Riesgo_Medio =("riesgo", lambda x: (x=="MEDIO 🟡").sum()),
        Vigentes     =("riesgo", lambda x: (x=="BAJO 🟢").sum()),
        Indefinidos  =("es_indefinido","sum"),
        Con_Garantia =("tiene_garantia","sum"),
        Monto        =("monto_total","sum") if "monto_total" in df.columns else ("id","count"),
    ).reset_index().sort_values("Contratos", ascending=False)
    resumen.rename(columns={"comprador_canon":"Comprador","tipo_comprador":"Tipo"}, inplace=True)
    if "Monto" in resumen.columns:
        resumen["Monto"] = resumen["Monto"].apply(fmt_m)
    st.dataframe(resumen, use_container_width=True, height=320)


# ══════════════════════════════════════════════
# TAB: PROVEEDORES
# ══════════════════════════════════════════════
with tab_prov:
    c1,c2 = st.columns(2)
    with c1:
        if "proveedor" in df.columns:
            top = df["proveedor"].value_counts().head(15).reset_index()
            top.columns = ["Proveedor","n"]; top["Proveedor"] = top["Proveedor"].str[:45]
            fig_p = px.bar(top, y="Proveedor", x="n", orientation="h",
                color="n", color_continuous_scale=[[0,"#B8D4EF"],[1,"#005CA9"]],
                labels={"n":"","Proveedor":""})
            fig_p.update_traces(marker_line_width=0)
            fig_p.update_layout(
                yaxis=dict(categoryorder="total ascending",tickfont=dict(size=9)),
                xaxis=dict(gridcolor="#F0F4F8"), coloraxis_showscale=False,
                paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
                font=dict(family="Inter",size=10),
                title=dict(text="Top 15 proveedores",
                           font=dict(size=12,color="#1A2E44",family="Inter"),x=0.01),
                height=380, margin=dict(t=38,b=10))
            st.plotly_chart(fig_p, use_container_width=True)
    with c2:
        if "proveedor" in df.columns:
            pv = df[df["riesgo"]=="ALTO 🔴"].groupby("proveedor").size().reset_index(name="n")
            pv = pv.sort_values("n", ascending=False).head(15)
            pv["proveedor"] = pv["proveedor"].str[:45]
            if not pv.empty:
                fig_pv = px.bar(pv, y="proveedor", x="n", orientation="h",
                    color="n", color_continuous_scale=[[0,"#FCA5A5"],[1,"#E02020"]],
                    labels={"proveedor":"","n":""})
                fig_pv.update_traces(marker_line_width=0)
                fig_pv.update_layout(
                    yaxis=dict(categoryorder="total ascending",tickfont=dict(size=9)),
                    xaxis=dict(gridcolor="#F0F4F8"), coloraxis_showscale=False,
                    paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
                    font=dict(family="Inter",size=10),
                    title=dict(text="Proveedores con más contratos vencidos",
                               font=dict(size=12,color="#1A2E44",family="Inter"),x=0.01),
                    height=380, margin=dict(t=38,b=10))
                st.plotly_chart(fig_pv, use_container_width=True)
            else:
                st.success("✅ Ningún proveedor tiene contratos en riesgo ALTO.")


# ══════════════════════════════════════════════
# TAB: GARANTÍAS
# ══════════════════════════════════════════════
with tab_gar_tab:
    c1,c2 = st.columns(2)
    with c1:
        gc = df["tiene_garantia"].map(
            {True:"Con garantía ✅",False:"Sin garantía ❌"}).value_counts()
        fig_g = go.Figure(go.Pie(
            labels=gc.index, values=gc.values, hole=0.55,
            marker_colors=["#00A651","#D1DCE8"],
            marker_line=dict(color="#ffffff", width=2),
            textinfo="percent+value", textfont=dict(size=11,family="Inter")))
        fig_g.update_layout(
            title=dict(text="Aplicación de garantías",
                       font=dict(size=12,color="#1A2E44",family="Inter"),x=0.02),
            paper_bgcolor="#ffffff", font=dict(family="Inter",size=10),
            height=250, margin=dict(t=38,b=10))
        st.plotly_chart(fig_g, use_container_width=True)
    with c2:
        gc2 = df[df["tiene_garantia"]].groupby("comprador_canon").size().reset_index(name="n")
        if not gc2.empty:
            fig_gc2 = px.bar(gc2.sort_values("n"), y="comprador_canon", x="n",
                orientation="h", color="n",
                color_continuous_scale=[[0,"#A7F3D0"],[1,"#00A651"]],
                labels={"comprador_canon":"","n":""})
            fig_gc2.update_traces(marker_line_width=0)
            fig_gc2.update_layout(
                yaxis=dict(categoryorder="total ascending",tickfont=dict(size=9)),
                xaxis=dict(gridcolor="#F0F4F8"), coloraxis_showscale=False,
                paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
                font=dict(family="Inter",size=10),
                title=dict(text="Garantías por comprador",
                           font=dict(size=12,color="#1A2E44",family="Inter"),x=0.01),
                height=270, margin=dict(t=38,b=10))
            st.plotly_chart(fig_gc2, use_container_width=True)

    df_grisk = df[df["tiene_garantia"] & df["riesgo"].isin(["ALTO 🔴","MEDIO 🟡"])].copy()
    if not df_grisk.empty:
        st.markdown("**⚠️ Contratos con garantía en riesgo:**")
        cols_gr = [c for c in ["id","proveedor","comprador_canon","estado_ariba",
                                "dias_venc","riesgo","fuente"] if c in df_grisk]
        st.dataframe(df_grisk[cols_gr].sort_values("dias_venc"),
                     use_container_width=True, height=200)


# ══════════════════════════════════════════════
# TAB: EXPLORADOR
# ══════════════════════════════════════════════
with tab_exp:
    cb,cn = st.columns([3,1])
    with cb:
        busq = st.text_input("🔎 Buscar proveedor, ID o descripción",
                             placeholder="Ej: LOGISTICA, CW2284016...")
    with cn:
        top_n = st.selectbox("Mostrar", [50,100,200,500,"Todos"], index=1)

    cols_def = [c for c in [
        "id","fuente","proveedor","comprador_canon","tipo_comprador","estado_ariba",
        "fecha_inicio","fecha_termino","dias_venc","riesgo","tiene_garantia",
        "monto_total","rut","area","gerencia"] if c in df.columns]
    cols_sel = st.multiselect("Columnas", df.columns.tolist(), default=cols_def)

    df_exp = df.copy()
    if busq.strip():
        mask = pd.Series(False, index=df_exp.index)
        for col in ["proveedor","proveedor_cons","descripcion","id","nombre_proyecto"]:
            if col in df_exp.columns:
                mask |= df_exp[col].astype(str).str.contains(
                    busq.strip(), case=False, na=False)
        df_exp = df_exp[mask]

    if "dias_venc" in df_exp.columns:
        df_exp = df_exp.sort_values("dias_venc")
    if top_n != "Todos":
        df_exp = df_exp.head(int(top_n))
    if cols_sel:
        df_exp = df_exp[cols_sel]

    def hl_fuente(val):
        if str(val) == "Solo Consolidado": return "background:#F5F0FF;color:#3B1A78;font-weight:600"
        if str(val) == "Ambos":            return "background:#F0FDF4;color:#065F26"
        if str(val) == "Ariba":            return "background:#EAF2FB;color:#005CA9"
        return ""

    if "fuente" in df_exp.columns:
        st.dataframe(df_exp.style.map(hl_fuente, subset=["fuente"]),
                     use_container_width=True, height=500)
    else:
        st.dataframe(df_exp, use_container_width=True, height=500)

    st.caption(
        f"Mostrando {len(df_exp):,} de {len(df):,} contratos "
        f"· {len(df_universo):,} total en universo unificado")


# ──────────────────────────────────────────────────────────────
# EXPORTACIÓN GENERAL
# ──────────────────────────────────────────────────────────────
st.markdown('<div class="sec">📥 Exportar datos del universo unificado</div>',
            unsafe_allow_html=True)
ts = datetime.now().strftime("%Y%m%d_%H%M")
ex1,ex2,ex3,ex4,ex5 = st.columns(5)
with ex1:
    st.download_button("💾 Vista actual · CSV",
                       df.to_csv(index=False).encode("utf-8-sig"),
                       f"contratos_{ts}.csv", mime="text/csv")
with ex2:
    crit = df[df["riesgo"].isin(["ALTO 🔴","MEDIO 🟡"])]
    st.download_button("🔴 Solo en riesgo",
                       crit.to_csv(index=False).encode("utf-8-sig"),
                       f"contratos_riesgo_{ts}.csv", mime="text/csv")
with ex3:
    urg = (df[df["dias_venc"].between(0,60)] if "dias_venc" in df.columns
           else pd.DataFrame())
    st.download_button("⚠️ Vencen en 60 días",
                       urg.to_csv(index=False).encode("utf-8-sig"),
                       f"contratos_urgentes_{ts}.csv", mime="text/csv")
with ex4:
    gdf = df[df["tiene_garantia"]]
    st.download_button("🔒 Con garantía",
                       gdf.to_csv(index=False).encode("utf-8-sig"),
                       f"contratos_garantia_{ts}.csv", mime="text/csv")
with ex5:
    sc_exp = (df[df["fuente"] == "Solo Consolidado"] if "fuente" in df.columns
              else pd.DataFrame())
    st.download_button("📂 Solo Drive",
                       sc_exp.to_csv(index=False).encode("utf-8-sig"),
                       f"contratos_solo_drive_{ts}.csv", mime="text/csv")


# ──────────────────────────────────────────────────────────────
# DIAGNÓSTICO
# ──────────────────────────────────────────────────────────────
with st.expander("🔧 Diagnóstico técnico"):
    d1,d2,d3 = st.columns(3)
    with d1:
        fuente_counts = df_universo["fuente"].value_counts().to_dict()
        st.json({"Total Pivot (activos)":       len(df_piv),
                 "Total universo unificado":     len(df_universo),
                 "En ambas fuentes":             fuente_counts.get("Ambos",0),
                 "Solo en Ariba":                fuente_counts.get("Ariba",0),
                 "Solo en Drive":                fuente_counts.get("Solo Consolidado",0),
                 "Con garantía":                 int(df_universo["tiene_garantia"].sum()),
                 "Indefinidos":                  int(df_universo["es_indefinido"].sum()),
                 "Compradores oficiales":        int(df_universo["es_oficial"].sum())})
    with d2:
        st.json({"Contratos en vista":    len(df),
                 "Riesgo ALTO":           int((df["riesgo"]=="ALTO 🔴").sum()),
                 "Riesgo MEDIO":          int((df["riesgo"]=="MEDIO 🟡").sum()),
                 "Consolidado cargado":   df_cons_raw is not None,
                 "Actualizado":           datetime.now().strftime("%d/%m/%Y %H:%M")})
    with d3:
        meta_diag = cargar_metadata()
        od_diag   = _leer_od_config()
        st.json({
            "Persistencia activa":        os.path.exists(PERSIST_PIVOT),
            "Directorio":                 PERSIST_DIR,
            "Pivot en servidor":          meta_diag.get("pivot",{}).get("filename","—"),
            "Pivot subido":               meta_diag.get("pivot",{}).get("uploaded_at","—"),
            "Pivot MD5":                  (meta_diag.get("pivot",{}).get("hash_md5","—")[:12]+"…"
                                           if meta_diag.get("pivot",{}).get("hash_md5") else "—"),
            "Consolidado en servidor":    meta_diag.get("consolidado",{}).get("filename","—"),
            "OneDrive URL configurada":   bool(od_diag.get("url","")),
            "Último sync OneDrive":       od_diag.get("ultimo_sync_ok","—"),
        })


# ── Footer ────────────────────────────────────────────────────
od_footer = od_cfg.get("url","")
if os.path.exists(PERSIST_PIVOT) and od_footer:
    persist_label = "💾 Persistente · 🔗 OneDrive vinculado"
elif os.path.exists(PERSIST_PIVOT):
    persist_label = "💾 Datos persistentes en servidor"
else:
    persist_label = "⚡ Sin persistencia activa"

st.markdown(f"""
<div style="margin-top:32px;padding:14px 20px;
     background:linear-gradient(135deg,#003F7A,#005CA9);
     border-radius:12px;display:flex;justify-content:space-between;align-items:center;">
  <div style="color:rgba(232,242,251,0.9);font-size:.72rem;font-weight:500;">
    <span style="font-weight:800;color:#ffffff;">Softys Chile</span>
    &nbsp;·&nbsp; Compras Estratégicas
    &nbsp;·&nbsp; Fuentes: SAP Ariba + Consolidado Drive
    &nbsp;·&nbsp; <span style="color:#7DEFA3;">{persist_label}</span>
  </div>
  <div style="color:rgba(232,242,251,0.6);font-size:.68rem;">
    {datetime.now().strftime('%d/%m/%Y %H:%M')}
  </div>
</div>
""", unsafe_allow_html=True)